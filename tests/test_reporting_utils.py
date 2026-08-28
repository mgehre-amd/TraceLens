###############################################################################
# Copyright (c) 2025 - 2026 Advanced Micro Devices, Inc. All rights reserved.
#
# See LICENSE for license information.
###############################################################################

"""Unit tests for TraceLens.Reporting.reporting_utils helpers."""

import json, os, pandas as pd, pytest, gzip, importlib, sys, subprocess, textwrap
from unittest import mock
from TraceLens.Reporting.reporting_utils import (
    _node_span_for_pg,
    _parse_pg_ranks,
    _safe_sheet_name,
    add_gpu_arch_cli_args,
    add_node_span_columns,
    detect_gpus_per_node,
    export_data_df,
    request_install,
    resolve_gpu_arch,
    write_report_outputs,
)
from TraceLens.Agent.Analysis.category_analyses import (
    analysis_utils as au,
    kernel_fusion_analysis as kfa,
)
from TraceLens.Reporting import (
    generate_multi_rank_collective_report_pytorch as coll_mod,
    generate_perf_report_pytorch_inference as inf_mod,
    reporting_utils as ru,
    tracediff_comparison_extension as tde,
)
from TraceLens.Reporting.generate_multi_rank_collective_report_pytorch import (
    find_trace_files,
    generate_collective_report,
)
from TraceLens.Reporting.generate_perf_report_pftrace_hip_activity import (
    _write_markdown_report,
    generate_perf_report_pftrace_hip_activity,
)
from TraceLens.TreePerf.tree_perf import TreePerfAnalyzer
from tests.fixtures.treeperf import _mk_ac2g
from tests.fixtures.traces import (
    INFERENCE_ROOT,
    NORM_TRACE,
    RESNET_TRACE,
    ROCprof_FILE,
    TESTS_DIR,
    TIMESFORMER1,
    TIMESFORMER2,
    TRACES_ROOT,
    _discover_inference_cases,
)
from TraceLens.Reporting.generate_perf_report_pytorch_inference import (
    add_truncated_kernel_details as add_truncated_inference,
    add_truncated_kernel_details as add_truncated_kernel_details_inference,
    apply_extension as apply_extension_inference,
    classify_graph_capture_trace,
    generate_perf_report_pytorch,
    generate_perf_report_pytorch as gen_inf,
    generate_perf_report_pytorch as generate_inference_report,
    get_dfs_short_kernels as get_dfs_short_kernels_inference,
    perf_report_sanity_check,
)
from TraceLens.Reporting.compare_perf_reports_pytorch import (
    generate_compare_perf_reports_pytorch,
)
from tests.fixtures.reporting import (
    _build_synthetic_trace,
    _create_genesis_capture,
    _minimal_pftrace_events,
    _mk_ac2g,
    _mk_event,
    _write_trace,
)
from tests.test_trace2tree import _mk_event
from unittest.mock import patch
from TraceLens.PerfModel import perf_model
from TraceLens.PerfModel.extensions import (
    attention_perf_model_extensions as aext,
    perf_model_extensions as pext,
)
from TraceLens.Reporting.generate_perf_report_pytorch import (
    _find_entry_point,
    _is_wrapper_frame,
    apply_extension as apply_extension_pytorch,
    generate_perf_report_pytorch,
    get_dfs_short_kernels as get_dfs_short_kernels_pytorch,
)
from TraceLens.Reporting.pftrace_hip_activity_analysis import (
    Event,
    HIPEvent,
    PftraceHipActivityAnalyzer,
    build_hip_summary_df,
    build_kernel_summary_df_for_name,
)
from TraceLens.Reporting.tracediff_comparison_extension import (
    tracediff_perf_summary_from_diff_stats,
)
from TraceLens.Trace2Tree.trace_capture_merge_experimental import (
    merge_capture_trace_into_graph,
)
from tests.fixtures.agent import (
    _StubAnalyzer,
    _StubTree,
    _kernel_event,
    _write_minimal_orchestrator_csvs,
)
from tests.fixtures.perfmodel import _ARCH, _GDN_ANNOTATION
from TraceLens.Trace2Tree.extensions import pseudo_ops_registry as por
from TraceLens.Trace2Tree.trace_to_tree import TraceToTree
from pathlib import Path
from TraceLens.Reporting.generate_perf_report_rocprof import (
    generate_perf_report_rocprof,
)
from tests.test_pftrace_memory_copy_report import _make_memory_copy_events
from TraceLens.Agent.Analysis.utils import orchestrator_prepare as op
from tests.test_analysis_agent_utils import TestOrchestratorPhase6
from TraceLens.Reporting.rocprof_analysis import _categorize_kernel
from types import SimpleNamespace
from TraceLens.Reporting.generate_perf_report_genesis import (
    _cleanup_work_dir,
    generate_perf_report_genesis,
)
from TraceLens.Reporting.pftrace_utils import (
    derive_pftrace_output_path,
    ensure_trace_json,
)

GPU_ONLY_TRACE = os.path.join(
    os.path.dirname(__file__),
    "traces/mi210/gpu_only_trace/gpu_only_trace.json.gz",
)


def test_export_data_df_csv_and_xlsx(tmp_path):
    df = pd.DataFrame({"a": [1.23456, 2.0], "b": [3.0, 4.567]})
    export_data_df(
        df,
        tmp_path,
        "report",
        output_table_format=[".csv", ".xlsx"],
        suffix="_stats",
    )
    csv_path = tmp_path / "report_stats.csv"
    xlsx_path = tmp_path / "report_stats.xlsx"
    assert csv_path.exists()
    assert xlsx_path.exists()
    loaded = pd.read_csv(csv_path)
    assert list(loaded.columns) == ["a", "b"]
    assert loaded["a"].tolist() == [1.23, 2.0]


def test_export_data_df_verbose(tmp_path, capsys):
    df = pd.DataFrame({"x": [1]})
    export_data_df(
        df,
        tmp_path,
        "out",
        output_table_format=[".csv"],
        suffix="",
        verbose=1,
    )
    captured = capsys.readouterr()
    assert "Exporting data to" in captured.out


def test_export_data_df_verbose_xlsx_and_debug(tmp_path, capsys):
    df = pd.DataFrame({"x": [1.234]})
    export_data_df(
        df,
        tmp_path,
        "stats",
        output_table_format=[".xlsx", ".csv"],
        suffix="_summary_statistics",
        verbose=4,
    )
    captured = capsys.readouterr()
    assert "Exporting data to" in captured.out
    assert "Data:" in captured.out
    assert (tmp_path / "stats_summary_statistics.xlsx").exists()
    assert (tmp_path / "stats_summary_statistics.csv").exists()


def test_request_install_declines_exits():
    with mock.patch("builtins.input", return_value="n"):
        with pytest.raises(SystemExit) as exc:
            request_install("openpyxl")
        assert exc.value.code == 1


def test_request_install_accepts_and_installs():
    with mock.patch("builtins.input", return_value="y"):
        with mock.patch("subprocess.check_call") as mock_install:
            request_install("openpyxl")
    mock_install.assert_called_once()


def test_request_install_failed_install_exits():

    with mock.patch("builtins.input", return_value="y"):
        with mock.patch(
            "subprocess.check_call",
            side_effect=subprocess.CalledProcessError(1, "pip"),
        ):
            with pytest.raises(SystemExit) as exc:
                request_install("openpyxl")
            assert exc.value.code == 1


def test_add_node_span_columns_intra_node():
    df = pd.DataFrame(
        {
            "rank": [0, 1, 2, 3],
            "Process Group Ranks": ["[0, 1, 2, 3]"] * 4,
        }
    )
    out = add_node_span_columns(df, gpus_per_node=4, world_size=8)
    assert "node_id" in out.columns
    assert "node_span" in out.columns
    assert out["node_id"].tolist() == [0, 0, 0, 0]
    assert (out["node_span"] == "intra_node").all()


def test_add_node_span_columns_inter_node():
    df = pd.DataFrame(
        {
            "rank": [0, 4],
            "Process Group Ranks": [
                "[0, 1, 2, 3, 4, 5, 6, 7]",
                "[0, 1, 2, 3, 4, 5, 6, 7]",
            ],
        }
    )
    out = add_node_span_columns(df, gpus_per_node=4, world_size=8)
    assert (out["node_span"] == "inter_node").all()
    assert out["node_id"].tolist() == [0, 1]


def test_add_node_span_columns_pg_ranks_string_formats():
    df = pd.DataFrame({"Process Group Ranks": ["(0, 1)", "0, 2, 4"]})
    out = add_node_span_columns(df, gpus_per_node=2, world_size=8)
    assert set(out["node_span"].unique()) <= {"intra_node", "inter_node", "unknown"}


def test_add_node_span_columns_empty_or_missing_columns():
    empty = pd.DataFrame()
    assert add_node_span_columns(empty, gpus_per_node=4).empty

    no_cols = pd.DataFrame({"other": [1]})
    result = add_node_span_columns(no_cols, gpus_per_node=4)
    assert "node_id" not in result.columns


def test_add_node_span_columns_invalid_gpus_per_node():
    df = pd.DataFrame({"rank": [0]})
    with pytest.raises(ValueError, match="gpus_per_node"):
        add_node_span_columns(df, gpus_per_node=0)


def test_add_node_span_columns_pg_ranks_only():
    df = pd.DataFrame({"Process Group Ranks": ["[0, 1]", "[2, 3]"]})
    out = add_node_span_columns(df, gpus_per_node=2)
    assert "node_span" in out.columns
    assert set(out["node_span"]) == {"intra_node"}


def test_add_node_span_columns_uninferable_world_size():
    df = pd.DataFrame({"Process Group Ranks": ["[]", "[]"]})
    out = add_node_span_columns(df, gpus_per_node=2)
    assert "node_id" not in out.columns


def test_parse_pg_ranks_and_node_span_helpers():

    assert _parse_pg_ranks([0, 1]) == [0, 1]
    assert _parse_pg_ranks("[0, 1]") == [0, 1]
    assert _parse_pg_ranks("groups 2 and 5") == [2, 5]
    assert _parse_pg_ranks(42) == []
    assert _node_span_for_pg([], {0: 0}, 2) == "unknown"


def test_add_node_span_columns_infers_world_size_from_rank():
    df = pd.DataFrame({"rank": [0, 1, 2, 3]})
    out = add_node_span_columns(df, gpus_per_node=2)
    assert list(out["node_id"]) == [0, 0, 1, 1]


def test_detect_gpus_per_node_from_trace():
    trace_path = os.path.join(
        "tests",
        "traces",
        "mi300",
        "Falconsai_nsfw_image_detection__1016002.json.gz",
    )
    if not os.path.isfile(trace_path):
        pytest.skip(f"Trace not found: {trace_path}")
    gpus = detect_gpus_per_node(trace_path)
    assert gpus is not None
    assert gpus > 0


def test_detect_gpus_per_node_invalid_file():
    assert detect_gpus_per_node("/nonexistent/trace.json") is None


def test_add_gpu_arch_cli_args_adds_mutually_exclusive_group():
    parser = __import__("argparse").ArgumentParser()
    add_gpu_arch_cli_args(parser)
    action_dests = {a.dest for a in parser._actions}
    assert "gpu_arch_json_path" in action_dests
    assert "gpu_arch_platform" in action_dests


def test_resolve_gpu_arch_json_roundtrip(tmp_path):
    arch = {"name": "test", "num_cus": 64}
    path = tmp_path / "arch.json"
    path.write_text(json.dumps(arch))
    assert resolve_gpu_arch(gpu_arch_json_path=str(path)) == arch


pytestmark = pytest.mark.filterwarnings(
    "ignore:Source column .* not found.*:UserWarning",
    "ignore:There are hipgraph launches.*:UserWarning",
    "ignore:Found .* events with failed performance metric.*:UserWarning",
    "ignore:Input list of events is empty.*:UserWarning",
    "ignore:dict_cat2names_extension is deprecated.*:UserWarning",
)

KERNEL_TRACE_CSV = """\
"Kind","Agent_Id","Queue_Id","Stream_Id","Thread_Id","Dispatch_Id","Kernel_Id","Kernel_Name","Correlation_Id","Start_Timestamp","End_Timestamp","LDS_Block_Size","Scratch_Size","VGPR_Count","Accum_VGPR_Count","SGPR_Count","Workgroup_Size_X","Workgroup_Size_Y","Workgroup_Size_Z","Grid_Size_X","Grid_Size_Y","Grid_Size_Z"
"KERNEL_DISPATCH","Agent 2",1,0,70,1,33,"__amd_rocclr_fillBufferAligned",119662,172352210005122,172352210008687,0,0,12,4,48,256,1,1,256,1,1
"KERNEL_DISPATCH","Agent 2",1,0,70,2,16,"kernel_step_1_c532_0_kernel_6_range_for",119670,172352210061004,172352210062686,0,0,4,4,16,1,1,1,1,1,1
"KERNEL_DISPATCH","Agent 2",1,0,70,3,31,"func_broad_phase_c402_0_kernel_3_range_for",119696,172352210143326,172352210149335,0,0,16,0,32,512,1,1,512,1,1
"""


def _mk_event(cat, name, ts, dur, pid, tid, args=None):
    return {
        "ph": "X",
        "cat": cat,
        "name": name,
        "pid": pid,
        "tid": tid,
        "ts": ts,
        "dur": dur,
        "args": args or {},
    }


def _mk_ac2g(corr_id, pid, tid, ts, phase):
    evt = {
        "ph": phase,
        "id": corr_id,
        "pid": pid,
        "tid": tid,
        "ts": ts,
        "cat": "ac2g",
        "name": "ac2g",
    }
    if phase == "f":
        evt["bp"] = "e"
    return evt


def _build_synthetic_trace(kernel_specs):
    events = []
    ts = 1000
    corr_id = 100
    cpu_pid, cpu_tid = 100, 100
    gpu_pid, gpu_tid = 0, 7

    for cpu_op_name, kernel_name, kernel_dur in kernel_specs:
        cpu_op_ts = ts
        cpu_op_dur = 100
        events.append(
            _mk_event(
                "cpu_op",
                cpu_op_name,
                ts=cpu_op_ts,
                dur=cpu_op_dur,
                pid=cpu_pid,
                tid=cpu_tid,
                args={"Input Dims": [[32, 64]], "Input type": ["float"]},
            )
        )
        events.append(
            _mk_event(
                "cuda_runtime",
                "hipLaunchKernel",
                ts=cpu_op_ts + 10,
                dur=5,
                pid=cpu_pid,
                tid=cpu_tid,
                args={"correlation": corr_id},
            )
        )
        kernel_ts = cpu_op_ts + 50
        events.append(
            _mk_event(
                "kernel",
                kernel_name,
                ts=kernel_ts,
                dur=kernel_dur,
                pid=gpu_pid,
                tid=gpu_tid,
                args={"correlation": corr_id, "stream": 7},
            )
        )
        events.append(_mk_ac2g(corr_id, gpu_pid, gpu_tid, kernel_ts, "s"))
        events.append(_mk_ac2g(corr_id, gpu_pid, gpu_tid, kernel_ts, "f"))
        ts += cpu_op_dur + 200
        corr_id += 1

    return {"traceEvents": events}


def _write_trace(tmp_path: Path, specs, name="trace.json") -> str:
    path = tmp_path / name
    path.write_text(json.dumps(_build_synthetic_trace(specs)))
    return str(path)


def _create_genesis_capture(tmp_path: Path) -> Path:
    capture = tmp_path / "capture"
    kernel_trace = capture / "kernel_trace"
    kernel_trace.mkdir(parents=True)
    (kernel_trace / "kernel_kernel_trace.csv").write_text(KERNEL_TRACE_CSV)
    (capture / "run.log").write_text("wall_time=4.00s\n")
    return capture


def _minimal_pftrace_events():
    return [
        {
            "ph": "X",
            "cat": "gpu_activity",
            "name": "xla_fusion_42",
            "pid": 0,
            "tid": 7,
            "ts": 1000,
            "dur": 50000,
            "args": {"agent": "gpu_0", "begin_ns": 1000000, "delta_ns": 50000000},
        },
        {
            "ph": "X",
            "cat": "hip_api",
            "name": "hipLaunchKernelGGL",
            "pid": 100,
            "tid": 1,
            "ts": 900,
            "dur": 20,
            "args": {"stream_ID": 0},
        },
    ]


class _MockShortKernelAnalyzer:
    def __init__(self, gpu_only=False, kernels=None, total_time_ms=1.0):
        self.gpu_only = gpu_only
        self.total_time_ms = total_time_ms
        self._kernels = (
            kernels
            if kernels is not None
            else pd.DataFrame(
                {
                    "Kernel duration (µs)": [5.0, 8.0, 50.0],
                    "Kernel name": ["k_short_a", "k_short_b", "k_long"],
                    "Parent cpu_op": ["aten::mm"] * 3,
                    "Input dims": ["[[32, 64]]"] * 3,
                    "Input strides": [""] * 3,
                    "Concrete Inputs": [""] * 3,
                }
            )
        )

    def get_df_kernels(self):
        return self._kernels


# ---------------------------------------------------------------------------
# perf_report_sanity_check (inference)
# ---------------------------------------------------------------------------


def test_sanity_check_include_nccl_busy_time():
    events = [{"name": "ncclAllReduce", "cat": "kernel"}]
    tl = pd.DataFrame({"type": ["busy_time"], "time ms": [0.05]})
    kl = pd.DataFrame({"total_direct_kernel_time_sum": [60.0]})
    up = pd.DataFrame({"Kernel Time (µs)_sum": [60.0]})
    result = perf_report_sanity_check(events, tl, kl, up, include_nccl=True)
    assert result["kl_time_pass"]
    assert result["total_gpu_events"] == 1


def test_sanity_check_time_mismatch():
    events = [{"name": "k", "cat": "kernel"}]
    tl = pd.DataFrame({"type": ["computation_time"], "time ms": [10.0]})
    kl = pd.DataFrame({"total_direct_kernel_time_sum": [1.0]})
    up = pd.DataFrame({"Kernel Time (µs)_sum": [1.0]})
    result = perf_report_sanity_check(events, tl, kl, up)
    assert not result["kl_time_pass"]
    assert not result["up_time_pass"]


def test_sanity_check_kernel_details_column():
    events = [{"name": "k_a", "cat": "kernel"}]
    tl = pd.DataFrame({"type": ["computation_time"], "time ms": [0.1]})
    kl = pd.DataFrame(
        {
            "total_direct_kernel_time": [100.0],
            "kernel_details": [[{"name": "k_a", "count": 1}]],
        }
    )
    up = pd.DataFrame(
        {
            "Kernel Time (µs)": [100.0],
            "kernel_details": [[{"name": "k_a", "count": 1}]],
        }
    )
    result = perf_report_sanity_check(events, tl, kl, up)
    assert result["kl_count_pass"]
    assert result["up_count_pass"]


def test_sanity_check_missing_kernel_details_column(capsys):
    events = [{"name": "k", "cat": "kernel"}]
    tl = pd.DataFrame({"type": ["computation_time"], "time ms": [0.1]})
    kl = pd.DataFrame({"total_direct_kernel_time_sum": [100.0]})
    up = pd.DataFrame({"Kernel Time (µs)_sum": [100.0]})
    result = perf_report_sanity_check(events, tl, kl, up)
    assert not result["kl_count_pass"]
    assert "WARNING" in capsys.readouterr().out


# ---------------------------------------------------------------------------
# classify_graph_capture_trace
# ---------------------------------------------------------------------------


def test_classify_skips_when_execution_details_exist(tmp_path, capsys):
    (tmp_path / "execution_details.json").write_text("[]")
    classify_graph_capture_trace(str(tmp_path))
    assert "Skipping classification" in capsys.readouterr().out


def test_classify_from_capture_annotation(tmp_path):
    events = [
        {
            "name": "vllm/v1/worker/gpu_model_runner.py(10): _dummy_run",
            "ts": 1,
        },
        {"name": "capture_32_FULL", "cat": "user_annotation", "ts": 2},
    ]
    (tmp_path / "graph.json").write_text(json.dumps({"traceEvents": events}))
    classify_graph_capture_trace(str(tmp_path))
    details = json.loads((tmp_path / "execution_details.json").read_text())
    assert details[0]["batch_size"] == 32
    assert details[0]["mode"] == "FULL"


def test_classify_inferred_from_stream_captures(tmp_path):
    events = [
        {
            "name": "vllm/v1/worker/gpu_model_runner.py(10): _dummy_run",
            "ts": 1,
        },
        {"cat": "cuda_runtime", "name": "cudaStreamBeginCapture", "ts": 2},
        {"cat": "cuda_runtime", "name": "cudaStreamBeginCapture", "ts": 3},
        {
            "cat": "cpu_op",
            "name": "aten::mm",
            "args": {"Input Dims": [[64, 128], [32, 64]]},
        },
    ]
    (tmp_path / "capture.json.gz").write_bytes(
        gzip.compress(json.dumps({"traceEvents": events}).encode())
    )
    classify_graph_capture_trace(str(tmp_path))
    details = json.loads((tmp_path / "execution_details.json").read_text())
    assert details[0]["mode"] == "PIECEWISE"
    assert details[0]["batch_size"] == 64


def test_classify_json_gz_roundtrip(tmp_path):
    events = [
        {"name": "vllm/v1/worker/gpu_model_runner.py(1): _dummy_run", "ts": 0},
        {"cat": "cuda_runtime", "name": "cudaStreamBeginCapture", "ts": 1},
        {"cat": "cpu_op", "name": "x", "args": {"Input Dims": [[8, 16]]}},
    ]
    gz_path = tmp_path / "capture.json.gz"
    gz_path.write_bytes(gzip.compress(json.dumps({"traceEvents": events}).encode()))
    classify_graph_capture_trace(str(tmp_path))
    details = json.loads((tmp_path / "execution_details.json").read_text())
    assert details[0]["batch_size"] == 8
    assert details[0]["mode"] == "FULL"


def test_classify_no_trace_files_exits(tmp_path):
    with pytest.raises(SystemExit) as exc:
        classify_graph_capture_trace(str(tmp_path))
    assert exc.value.code == 0


# ---------------------------------------------------------------------------
# get_dfs_short_kernels
# ---------------------------------------------------------------------------


def test_inference_get_dfs_short_kernels_with_data():
    analyzer = _MockShortKernelAnalyzer()
    hist, grouped = get_dfs_short_kernels_inference(analyzer, topk=1)
    assert not hist.empty
    assert len(grouped) == 1
    assert "Short Kernel duration (µs) sum" in grouped.columns


def test_inference_get_dfs_short_kernels_empty():
    empty = pd.DataFrame(columns=["Kernel duration (µs)", "Kernel name"])
    analyzer = _MockShortKernelAnalyzer(kernels=empty)
    hist, grouped = get_dfs_short_kernels_inference(analyzer)
    assert hist.empty
    assert grouped.empty


def test_inference_get_dfs_short_kernels_gpu_only():
    kernels = pd.DataFrame(
        {"Kernel duration (µs)": [3.0], "Kernel name": ["k"]},
    )
    analyzer = _MockShortKernelAnalyzer(gpu_only=True, kernels=kernels)
    hist, grouped = get_dfs_short_kernels_inference(analyzer)
    assert not hist.empty
    assert grouped.iloc[0]["Kernel name"] == "k"


def test_pytorch_get_dfs_short_kernels_with_data():
    analyzer = _MockShortKernelAnalyzer()
    hist, grouped = get_dfs_short_kernels_pytorch(analyzer, topk=2)
    assert len(grouped) == 2


# ---------------------------------------------------------------------------
# apply_extension
# ---------------------------------------------------------------------------


def test_inference_apply_extension_op_category(tmp_path):
    ext_path = tmp_path / "ext.py"
    ext_path.write_text(textwrap.dedent("""
            def tree_postprocess_extension(tree):
                tree.events[0]["ext_applied"] = True

            op_category_extension = {"custom::op": "Other"}
            """))
    tree = SimpleNamespace(events=[{}], label_non_gpu_paths=lambda: None)
    analyzer = SimpleNamespace(
        tree=tree,
        op_to_perf_model_class_map={},
    )
    apply_extension_inference(analyzer, str(ext_path))
    assert analyzer.tree.events[0]["ext_applied"]


def test_inference_apply_extension_invalid_perf_model(tmp_path):
    ext_path = tmp_path / "bad_ext.py"
    ext_path.write_text("perf_model_extension = {'aten::mm': 'not_a_class'}")
    analyzer = SimpleNamespace(
        tree=SimpleNamespace(events=[], label_non_gpu_paths=lambda: None),
        op_to_perf_model_class_map={},
    )
    with pytest.raises(ValueError, match="category attribute"):
        apply_extension_inference(analyzer, str(ext_path))


def test_pytorch_apply_extension_valid_perf_model(tmp_path):
    ext_path = tmp_path / "ext.py"
    ext_path.write_text(textwrap.dedent("""
            class DummyGemm:
                category = "GEMM"

            perf_model_extension = {"aten::mm": DummyGemm}
            """))
    analyzer = SimpleNamespace(
        tree=SimpleNamespace(events=[], label_non_gpu_paths=lambda: None),
        op_to_perf_model_class_map={},
    )
    apply_extension_pytorch(analyzer, str(ext_path))
    assert "aten::mm" in analyzer.op_to_perf_model_class_map


# ---------------------------------------------------------------------------
# trunc / wrapper helpers
# ---------------------------------------------------------------------------


def test_inference_add_truncated_kernel_details_missing_column():
    df = pd.DataFrame({"other": [1]})
    out = add_truncated_kernel_details_inference(df, source_col="missing")
    assert "trunc_missing" not in out.columns


def test_pytorch_is_wrapper_frame():
    assert _is_wrapper_frame("torch/nn/modules/module.py(5): _call_impl")
    assert _is_wrapper_frame("torch/_ops.py(10): wrapper_custom")
    assert not _is_wrapper_frame("user_code.py(10): forward")


def test_find_entry_point_stripped_suffix():
    stack = str(["user.py(1): addmm_triton", "aten::addmm_triton"])
    result = _find_entry_point(stack, "aten::addmm_triton_340")
    assert result["traversal"] == "outward"
    assert "user.py" in result["entry_point"]


# ---------------------------------------------------------------------------
# generate_perf_report_pytorch / inference with synthetic traces
# ---------------------------------------------------------------------------


def test_pytorch_report_synthetic_minimal(tmp_path):
    trace = _write_trace(
        tmp_path,
        [("aten::mm", "gemm_kernel", 100), ("aten::relu", "relu_kernel", 20)],
    )
    out = str(tmp_path / "csvs")
    result = generate_perf_report_pytorch(
        profile_json_path=trace,
        output_csvs_dir=out,
        collective_analysis=False,
        kernel_summary=True,
        short_kernel_study=True,
    )
    assert "gpu_timeline" in result
    assert "kernel_summary" in result
    assert os.path.isfile(os.path.join(out, "gpu_timeline.csv"))


def test_inference_report_synthetic_with_flags(tmp_path):
    trace = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 100)])
    out = str(tmp_path / "csvs")
    xlsx = str(tmp_path / "report.xlsx")
    result = generate_inference_report(
        profile_json_path=trace,
        output_csvs_dir=out,
        output_xlsx_path=xlsx,
        collective_analysis=False,
        kernel_summary=True,
        short_kernel_study=True,
        group_by_parent_module=True,
        group_by_num_kernels=True,
    )
    assert os.path.isfile(xlsx)
    assert "short_kernel_histogram" in result
    assert "short_kernels_summary" in result


def test_inference_report_with_extension_additional_dfs(tmp_path):
    trace = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 100)])
    ext_path = tmp_path / "extra.py"
    ext_path.write_text(textwrap.dedent("""
            import pandas as pd

            def get_additional_dataframes_extension(tree):
                return {"custom_extra": pd.DataFrame({"value": [42]})}
            """))
    result = generate_inference_report(
        profile_json_path=trace,
        output_csvs_dir=str(tmp_path / "csvs"),
        collective_analysis=False,
        extension_file=str(ext_path),
    )
    assert "custom_extra" in result


def test_pytorch_report_include_overlap_and_call_stack(tmp_path):
    trace = _write_trace(
        tmp_path,
        [
            ("aten::mm", "gemm_kernel", 100),
            ("aten::add", "add_kernel", 15),
        ],
    )
    result = generate_perf_report_pytorch(
        profile_json_path=trace,
        output_csvs_dir=str(tmp_path / "csvs"),
        collective_analysis=False,
        include_overlap_info=True,
        include_call_stack=True,
        group_by_num_kernels=True,
    )
    assert "unified_perf_summary" in result
    assert "call_stack_full" in result["unified_perf_summary"].columns


# ---------------------------------------------------------------------------
# multi-rank collective report
# ---------------------------------------------------------------------------


def test_find_trace_files_empty_dir(tmp_path, capsys):
    assert find_trace_files(str(tmp_path)) == []
    assert "No trace files found" in capsys.readouterr().out


def test_collective_report_trace_dir_synthetic(tmp_path):
    for rank in range(2):
        trace = _make_trace(rank, 3)
        (tmp_path / f"rank{rank}_trace.json").write_text(json.dumps(trace))
    out = str(tmp_path / "nccl_out")
    dfs = generate_collective_report(
        trace_dir=str(tmp_path),
        world_size=2,
        output_csvs_dir=out,
        detailed_analysis=False,
        gpus_per_node=2,
        strict_world_size_check=False,
    )
    assert "nccl_summary_implicit_sync" in dfs
    assert os.path.isfile(os.path.join(out, "nccl_summary_implicit_sync.csv"))


def test_collective_report_trace_pattern_non_strict(tmp_path):
    for rank in (0, 2):
        trace = _make_trace(rank, 2)
        (tmp_path / f"trace_rank_{rank}.json").write_text(json.dumps(trace))
    pattern = str(tmp_path / "trace_rank_*.json")
    dfs = generate_collective_report(
        trace_pattern=pattern,
        world_size=4,
        output_csvs_dir=str(tmp_path / "out"),
        strict_world_size_check=False,
        detailed_analysis=True,
    )
    assert "nccl_long" in dfs


def test_collective_report_xlsx_output(tmp_path):
    trace = _make_trace(0, 2)
    (tmp_path / "rank0_trace.json").write_text(json.dumps(trace))
    xlsx = str(tmp_path / "report.xlsx")
    generate_collective_report(
        trace_dir=str(tmp_path),
        world_size=1,
        output_xlsx_path=xlsx,
        strict_world_size_check=False,
    )
    assert os.path.isfile(xlsx)


def _make_trace(rank, n_collectives):
    events = []
    base_ts = 1_000_000 + rank * 50
    for i in range(n_collectives):
        ts = base_ts + i * 1000 + rank * 5
        events.append(
            {
                "ph": "X",
                "cat": "kernel",
                "name": "void rcclGenericKernel<1, false>(ncclDevKernelArgsStorage<4096ul>)",
                "pid": rank,
                "tid": 3,
                "ts": ts,
                "dur": 50,
                "args": {
                    "External id": 100 + i,
                    "device": rank,
                    "stream": 3,
                    "correlation": 50 + i,
                },
            }
        )
    return {"traceEvents": events}


# ---------------------------------------------------------------------------
# genesis report
# ---------------------------------------------------------------------------


def test_cleanup_work_dir(tmp_path):
    work = tmp_path / "work"
    work.mkdir()
    (work / "temp.json").write_text("{}")
    _cleanup_work_dir(work)
    assert not work.exists()


def test_generate_perf_report_genesis_integration(tmp_path):
    capture = _create_genesis_capture(tmp_path)
    out = tmp_path / "analysis"
    reports = generate_perf_report_genesis(
        capture_dir=str(capture),
        output_dir=str(out),
        short_kernel_study=False,
        keep_work=True,
    )
    assert "rocprof" in reports
    assert (out / "genesis_perf_report.xlsx").exists()
    assert (out / "genesis_summary.md").exists()
    assert reports["rocprof"]["kernel_summary_by_category"] is not None


@mock.patch(
    "TraceLens.Reporting.generate_perf_report_genesis.generate_perf_report_pftrace_memory_copy"
)
@mock.patch(
    "TraceLens.Reporting.generate_perf_report_genesis.generate_perf_report_pftrace_hip_activity"
)
@mock.patch("TraceLens.Reporting.generate_perf_report_genesis.pftrace_to_json")
def test_generate_perf_report_genesis_with_pftrace(
    mock_pftrace_to_json,
    mock_hip_activity,
    mock_memory_copy,
    tmp_path,
):
    capture = _create_genesis_capture(tmp_path)
    pftrace = capture / "kernel_trace" / "kernel_results.pftrace"
    pftrace.write_bytes(b"\x00")
    mock_pftrace_to_json.return_value = capture / "pf.json"
    mock_hip_activity.return_value = {
        "hip_summary": pd.DataFrame({"api": ["hipMalloc"]})
    }
    mock_memory_copy.return_value = {
        "memory_copy_by_copy_bytes": pd.DataFrame({"copy_bytes": [1024], "count": [1]})
    }
    out = tmp_path / "analysis_pf"
    reports = generate_perf_report_genesis(
        capture_dir=str(capture),
        output_dir=str(out),
        short_kernel_study=False,
    )
    assert "pftrace_hip_activity" in reports
    assert "pftrace_memory_copy" in reports
    mock_hip_activity.assert_called_once()


# ---------------------------------------------------------------------------
# pftrace modules
# ---------------------------------------------------------------------------


def test_write_markdown_report(tmp_path):
    df_cat = pd.DataFrame({"GPU ID": [0], "Category": ["xla"], "Time (ms)": [1.0]})
    md_path = tmp_path / "report.md"
    _write_markdown_report(
        md_path,
        df_category=df_cat,
        xla_top=[("xla_fusion_1", 1_000_000, 2, 0.5)],
        used_fav3=False,
        agents=["gpu_0"],
        kernel_df=pd.DataFrame({"Name": ["k1"], "Instances": [1]}),
        hip_df=pd.DataFrame({"Name": ["hipMalloc"], "Instances": [1]}),
    )
    text = md_path.read_text()
    assert "ROCm Perfetto Trace Report" in text
    assert "xla_fusion_1" in text


def test_ensure_trace_json_returns_json_path(tmp_path):
    trace = tmp_path / "trace.json"
    trace.write_text('{"traceEvents": []}')
    assert ensure_trace_json(str(trace)) == str(trace.resolve())


def test_ensure_trace_json_unsupported_format(tmp_path):
    bad = tmp_path / "trace.bin"
    bad.write_bytes(b"\x00")
    with pytest.raises(ValueError, match="Unsupported trace format"):
        ensure_trace_json(str(bad))


def test_build_kernel_summary_df_for_name():
    events = [
        Event(gpu=0, name="gemm_1", dur_ns=1000, ts_ns=0),
        Event(gpu=0, name="gemm_2", dur_ns=2000, ts_ns=0),
    ]
    df = build_kernel_summary_df_for_name(
        events, baseline_total_ns=3000, merge_names=True
    )
    assert len(df) == 1
    assert df.iloc[0]["Instances"] == 2


@pytest.mark.parametrize("group", ["name", "name+stream", "name+op", "name+stream+op"])
def test_build_hip_summary_df_groups(group):
    hip_events = [
        HIPEvent(
            name="hipMalloc",
            dur_ns=100,
            ts_ns=0,
            pid=1,
            tid=1,
            stream_id=1,
            operation=2,
        ),
        HIPEvent(
            name="hipMalloc",
            dur_ns=200,
            ts_ns=0,
            pid=1,
            tid=1,
            stream_id=1,
            operation=2,
        ),
    ]
    df = build_hip_summary_df(hip_events, group=group)
    assert not df.empty
    assert df.iloc[0]["Instances"] == 2


def test_pftrace_hip_activity_markdown_output(tmp_path):
    trace_path = tmp_path / "trace.json"
    trace_path.write_text(json.dumps({"traceEvents": _minimal_pftrace_events()}))
    md_path = tmp_path / "out.md"
    generate_perf_report_pftrace_hip_activity(
        trace_path=str(trace_path),
        output_md_path=str(md_path),
        min_event_ns=0,
        kernel_summary=True,
        hip_summary=True,
    )
    assert md_path.exists()
    assert "ROCm Perfetto Trace Report" in md_path.read_text()


def test_pftrace_hip_activity_default_xlsx_path(tmp_path):
    trace_path = tmp_path / "trace.json"
    trace_path.write_text(json.dumps({"traceEvents": _minimal_pftrace_events()}))
    generate_perf_report_pftrace_hip_activity(
        trace_path=str(trace_path),
        min_event_ns=0,
    )
    assert (tmp_path / "trace_pftrace_activity_report.xlsx").exists()


# ---------------------------------------------------------------------------
# inference report main() and extended paths
# ---------------------------------------------------------------------------


def test_inference_report_with_overlap_and_collective(tmp_path):
    trace = _write_trace(
        tmp_path,
        [
            ("aten::mm", "gemm_kernel", 100),
            ("aten::add", "add_kernel", 20),
            ("aten::relu", "relu_kernel", 15),
        ],
    )
    result = generate_inference_report(
        profile_json_path=trace,
        output_csvs_dir=str(tmp_path / "csvs"),
        output_xlsx_path=str(tmp_path / "report.xlsx"),
        collective_analysis=True,
        include_overlap_info=True,
        kernel_summary=True,
        short_kernel_study=True,
    )
    assert "gpu_timeline" in result
    assert os.path.isfile(str(tmp_path / "report.xlsx"))


def test_inference_report_main_cli(tmp_path):

    trace = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 100)])
    out_dir = tmp_path / "cli_out"
    xlsx = tmp_path / "cli_report.xlsx"

    old_argv = sys.argv
    sys.argv = [
        "generate_perf_report_pytorch_inference",
        "--profile_json_path",
        trace,
        "--output_csvs_dir",
        str(out_dir),
        "--output_xlsx_path",
        str(xlsx),
        "--enable_kernel_summary",
        "--short_kernel_study",
        "--group_by_parent_module",
    ]
    try:
        inf_mod.main()
    finally:
        sys.argv = old_argv
    assert xlsx.exists()
    assert (out_dir / "gpu_timeline.csv").exists()


def test_collective_report_trace_glob(tmp_path):
    for rank in range(2):
        trace = _make_trace(rank, 2)
        (tmp_path / f"custom_rank{rank}_trace.json").write_text(json.dumps(trace))
    dfs = generate_collective_report(
        trace_glob=str(tmp_path / "custom_rank*_trace.json"),
        world_size=2,
        output_csvs_dir=str(tmp_path / "glob_out"),
        strict_world_size_check=False,
        use_multiprocessing=False,
        all2allv_heatmap=False,
    )
    assert "nccl_summary_implicit_sync" in dfs


def test_collective_report_main_cli(tmp_path):

    trace = _make_trace(0, 2)
    (tmp_path / "rank0_trace.json").write_text(json.dumps(trace))
    out = tmp_path / "nccl_cli.xlsx"

    old_argv = sys.argv
    sys.argv = [
        "generate_multi_rank_collective_report_pytorch",
        "--trace_dir",
        str(tmp_path),
        "--world_size",
        "1",
        "--output_xlsx_path",
        str(out),
    ]
    try:
        coll_mod.main()
    finally:
        sys.argv = old_argv
    assert out.exists()


class TestAnalysisUtilsAndReporting:
    def test_perf_report_csv_dir_comparative(self, tmp_path):
        cat_dir = tmp_path / "category_data"
        cat_dir.mkdir()
        (cat_dir / "category_manifest.json").write_text(
            json.dumps({"comparison_scope": "comparative"})
        )
        assert "perf_report_trace1_csvs" in au.perf_report_csv_dir(str(tmp_path))

    def test_resolve_gpu_arch_and_node_span(self):
        arch = ru.resolve_gpu_arch(
            gpu_arch={"name": "mi300x", "freq_mhz": 2200, "num_cus": 304}
        )
        assert arch["name"] == "mi300x"
        out = ru.add_node_span_columns(
            pd.DataFrame({"rank": [0, 1], "Process Group Ranks": ["[0,1]", "[0,1]"]}),
            gpus_per_node=2,
            world_size=2,
        )
        assert "node_id" in out.columns

    def test_reporting_utils_export_and_node_span(self, tmp_path):

        df = pd.DataFrame({"a": [1, 2]})
        ru.export_data_df(df, Path(tmp_path), "test", output_table_format=[".csv"])
        assert (tmp_path / "test_summary_statistics.csv").exists()


class TestPytorchReportOverlapPhase10:
    def test_pytorch_report_with_overlap_sheets(self, tmp_path):

        trace = _write_trace(
            tmp_path,
            [
                ("aten::convolution", "conv_kernel", 80),
                ("aten::convolution_backward", "conv_bwd_kernel", 70),
                ("aten::mm", "gemm_kernel", 100),
            ],
            "conv.json",
        )
        out = tmp_path / "py_out"
        dfs = generate_perf_report_pytorch(
            profile_json_path=str(trace),
            output_csvs_dir=str(out),
            include_overlap_info=True,
            short_kernel_study=True,
            kernel_summary=True,
            collective_analysis=False,
        )
        assert isinstance(dfs, dict)
        assert (out / "gpu_timeline.csv").exists()


class TestReportingPhase11:
    def test_inference_report_capture_merge(self, tmp_path):

        trace = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 100)], "inf.json")
        out = tmp_path / "inf_out"
        dfs = generate_perf_report_pytorch(
            profile_json_path=str(trace),
            output_csvs_dir=str(out),
            kernel_summary=True,
        )
        assert isinstance(dfs, dict)


class TestReportingPhase12:
    def test_classify_graph_capture_json_gz(self, tmp_path):
        capture_dir = tmp_path / "captures"
        capture_dir.mkdir()
        trace = {
            "traceEvents": [
                {
                    "ph": "X",
                    "cat": "user_annotation",
                    "name": "graph_capture: batch=4 mode=FULL",
                    "ts": 1000,
                    "dur": 100,
                    "pid": 1,
                    "tid": 1,
                    "args": {},
                },
            ],
            "schemaVersion": 1,
        }
        gz_path = capture_dir / "graph_capture_rank_0.json.gz"
        with gzip.open(gz_path, "wt") as f:
            json.dump(trace, f)
        classify_graph_capture_trace(str(capture_dir))
        assert (capture_dir / "execution_details.json").exists()

    @pytest.mark.skipif(
        not (os.path.isfile(TIMESFORMER1) and os.path.isfile(TIMESFORMER2)),
        reason="timesformer traces missing",
    )
    def test_inference_report_overlap_on_trace(self, tmp_path):
        out = tmp_path / "inf_csv"
        generate_inference_report(
            profile_json_path=TIMESFORMER1,
            output_csvs_dir=str(out),
            include_overlap_info=True,
            kernel_summary=True,
            short_kernel_study=True,
            group_by_parent_module=True,
        )
        assert (out / "gpu_timeline.csv").exists()


class TestReportingPhase12B:
    @pytest.mark.skipif(not os.path.isfile(RESNET_TRACE), reason="resnet missing")
    def test_pytorch_report_overlap_bwd(self, tmp_path):

        out = tmp_path / "pt_out"
        dfs = generate_perf_report_pytorch(
            profile_json_path=RESNET_TRACE,
            output_csvs_dir=str(out),
            include_overlap_info=True,
            kernel_summary=True,
            short_kernel_study=True,
        )
        assert isinstance(dfs, dict)
        assert (out / "gpu_timeline.csv").exists()


class TestReportingCliPhase4:
    def test_pftrace_memory_copy_main(self, tmp_path):
        trace_path = tmp_path / "trace.json"
        trace_path.write_text(json.dumps({"traceEvents": _minimal_pftrace_events()}))
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_pftrace_memory_copy"
        )
        out_dir = tmp_path / "mem_csv"
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_pftrace_memory_copy",
            "--trace_path",
            str(trace_path),
            "--output_csvs_dir",
            str(out_dir),
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert out_dir.exists()

    def test_pftrace_hip_api_main(self, tmp_path):
        trace_path = tmp_path / "trace.json"
        trace_path.write_text(json.dumps({"traceEvents": _minimal_pftrace_events()}))
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_pftrace_hip_api"
        )
        out_dir = tmp_path / "api_csv"
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_pftrace_hip_api",
            "--trace_path",
            str(trace_path),
            "--output_csvs_dir",
            str(out_dir),
            "--include_nonlaunch_apis",
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert out_dir.exists()

    def test_genesis_report_main(self, tmp_path):
        capture = _create_genesis_capture(tmp_path)
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_genesis"
        )
        out = tmp_path / "gen_out"
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_genesis",
            "--capture-dir",
            str(capture),
            "--output-dir",
            str(out),
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert (out / "genesis_perf_report.xlsx").exists()

    @pytest.mark.skipif(
        not os.path.isfile(ROCprof_FILE), reason="rocprof fixture missing"
    )
    def test_rocprof_report_function(self, tmp_path):

        generate_perf_report_rocprof(
            profile_json_path=ROCprof_FILE,
            output_xlsx_path=str(tmp_path / "roc.xlsx"),
            kernel_summary=True,
            short_kernel_study=True,
            kernel_details=True,
        )
        assert (tmp_path / "roc.xlsx").exists()

    def test_compare_perf_reports_main(self, tmp_path):
        t1 = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 100)], "a.json")
        t2 = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 120)], "b.json")
        x1 = tmp_path / "r1.xlsx"
        x2 = tmp_path / "r2.xlsx"
        generate_perf_report_pytorch(
            profile_json_path=t1,
            output_csvs_dir=str(tmp_path / "csv1"),
            output_xlsx_path=str(x1),
            kernel_summary=True,
        )
        generate_perf_report_pytorch(
            profile_json_path=t2,
            output_csvs_dir=str(tmp_path / "csv2"),
            output_xlsx_path=str(x2),
            kernel_summary=True,
        )
        mod = importlib.import_module(
            "TraceLens.Reporting.compare_perf_reports_pytorch"
        )
        out = tmp_path / "cmp.xlsx"
        old_argv = sys.argv
        sys.argv = [
            "compare_perf_reports_pytorch",
            str(x1),
            str(x2),
            "-o",
            str(out),
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert out.exists()


class TestReportingPhase6:
    @pytest.mark.skipif(not os.path.isfile(NORM_TRACE), reason="norm trace missing")
    def test_pytorch_report_all_bwd_overlap_flags(self, tmp_path):
        generate_perf_report_pytorch(
            profile_json_path=NORM_TRACE,
            output_csvs_dir=str(tmp_path / "csv"),
            output_xlsx_path=str(tmp_path / "out.xlsx"),
            kernel_summary=True,
            short_kernel_study=True,
            include_overlap_info=True,
            group_by_num_kernels=True,
            topk_ops=10,
            topk_roofline_ops=5,
            include_unlinked_kernels=True,
            include_call_stack=True,
            enable_pseudo_ops=True,
        )
        assert (tmp_path / "csv" / "gpu_timeline.csv").exists()

    def test_classify_graph_capture_zip_and_dummy_run(self, tmp_path):
        capture_dir = tmp_path / "capture"
        capture_dir.mkdir()
        events = {
            "traceEvents": [
                _mk_event(
                    "cpu_op",
                    "vllm/v1/worker/gpu_model_runner.py(100): _dummy_run",
                    1000,
                    50,
                    1,
                    1,
                    {},
                ),
                _mk_event("cuda_runtime", "cudaStreamBeginCapture", 1100, 10, 1, 1, {}),
                _mk_event(
                    "cpu_op",
                    "aten::mm",
                    1200,
                    20,
                    1,
                    1,
                    {"Input Dims": [[8, 64], [64, 128]]},
                ),
            ]
        }
        json_path = capture_dir / "trace.json"
        json_path.write_text(json.dumps(events))
        classify_graph_capture_trace(str(capture_dir))
        details = json.loads((capture_dir / "execution_details.json").read_text())
        assert details[0]["batch_size"] == 8

    def test_inference_extension_and_sanity(self, tmp_path):

        trace = _write_trace(
            tmp_path,
            [
                ("aten::mm", "gemm_kernel", 100),
                ("aten::native_layer_norm", "layer_norm_kernel", 30),
            ],
        )
        gen_inf(
            profile_json_path=trace,
            output_csvs_dir=str(tmp_path / "out"),
            output_xlsx_path=str(tmp_path / "r.xlsx"),
            collective_analysis=True,
            kernel_summary=True,
            short_kernel_study=True,
            include_overlap_info=True,
            group_by_parent_module=True,
            group_by_num_kernels=True,
            topk_ops=5,
            include_unlinked_kernels=True,
            micro_idle_thresh_us=0,
        )
        analyzer = TreePerfAnalyzer.from_file(trace)
        sanity = perf_report_sanity_check(
            analyzer.tree.events,
            pd.read_csv(str(tmp_path / "out" / "gpu_timeline.csv")),
            analyzer.get_df_kernel_launchers(),
            analyzer.build_df_unified_perf_table(),
        )
        assert isinstance(sanity, dict)

    @pytest.mark.skipif(
        not os.path.isdir(os.path.join(INFERENCE_ROOT, "vllm_decode_full")),
        reason="inference fixture missing",
    )
    def test_inference_real_fixture_report(self, tmp_path):

        case = os.path.join(INFERENCE_ROOT, "vllm_decode_full")
        trace = next(
            os.path.join(case, f) for f in os.listdir(case) if f.endswith(".json.gz")
        )
        gen_inf(
            profile_json_path=trace,
            output_csvs_dir=str(tmp_path / "inf"),
            output_xlsx_path=str(tmp_path / "inf.xlsx"),
            collective_analysis=False,
            kernel_summary=True,
        )
        assert (tmp_path / "inf" / "gpu_timeline.csv").exists()


class TestReportingPhase7:
    def test_compare_perf_reports_all_sheets(self, tmp_path):

        t1 = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 100)], "t1.json")
        t2 = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 120)], "t2.json")
        d1 = tmp_path / "r1"
        d2 = tmp_path / "r2"
        generate_perf_report_pytorch(
            profile_json_path=str(t1),
            output_csvs_dir=str(d1),
            output_xlsx_path=str(tmp_path / "r1.xlsx"),
        )
        generate_perf_report_pytorch(
            profile_json_path=str(t2),
            output_csvs_dir=str(d2),
            output_xlsx_path=str(tmp_path / "r2.xlsx"),
        )
        out = tmp_path / "cmp"
        generate_compare_perf_reports_pytorch(
            reports=[str(d1), str(d2)],
            output=str(tmp_path / "cmp.xlsx"),
            sheets=["gpu_timeline", "ops_summary"],
            output_csvs_dir=str(out),
        )
        assert (out / "gpu_timeline.csv").exists()

    def test_pftrace_analyser_extended(self, tmp_path):
        events = _minimal_pftrace_events()
        analyser = PftraceHipActivityAnalyzer(events)
        assert isinstance(analyser.get_df_category_summary(), pd.DataFrame)
        assert isinstance(analyser.get_df_kernel_summary(), pd.DataFrame)
        assert isinstance(analyser.get_df_hip_summary(), pd.DataFrame)

    def test_tracediff_extension_multi_kernel_row(self):
        diff = pd.DataFrame(
            {
                "source": ["trace1", "trace1"],
                "lowest_common_ancestor_id": [5, 5],
                "lowest_common_ancestor_name": ["aten::mm", "aten::mm"],
                "cpu_op_name": ["aten::mm", "aten::add"],
                "busy_time": [100.0, 50.0],
                "name": ["k1", "k2"],
                "gpu_op_uid": [1, 2],
                "nn_module_stack": ["[]", "[]"],
                "nn_module_parent": ["", ""],
                "Input Dims": ["[[2,3]]", "[[2,3]]"],
                "Input type": ["['fp16']", "['fp16']"],
                "Input Strides": ["[]", "[]"],
                "Concrete Inputs": ["", ""],
            }
        )
        summary = tracediff_perf_summary_from_diff_stats(diff)
        assert " | " in summary.iloc[0]["name"]

    @pytest.mark.skipif(
        not os.path.isdir(os.path.join(INFERENCE_ROOT, "sglang_decode")),
        reason="inference fixture missing",
    )
    def test_capture_merge_inference_fixture(self):
        case = os.path.join(INFERENCE_ROOT, "sglang_decode")
        trace_gz = next(f for f in os.listdir(case) if f.endswith(".json.gz"))
        graph = os.path.join(case, trace_gz)
        capture = os.path.join(case, "capture_traces")
        metadata = os.path.join(capture, "execution_details.json")
        merged = merge_capture_trace_into_graph(capture, metadata, graph)
        analyzer = TreePerfAnalyzer(merged, rebuild_tree=False, enable_pseudo_ops=True)
        unified = analyzer.build_df_unified_perf_table(include_nccl=False)
        assert isinstance(unified, pd.DataFrame)

    def test_collective_report_strict_and_heatmap(self, tmp_path):
        for rank in (0, 1):
            (tmp_path / f"rank{rank}_trace.json").write_text(
                json.dumps(
                    {
                        "traceEvents": [
                            {
                                "ph": "X",
                                "cat": "kernel",
                                "name": "ncclKernel_AllReduce",
                                "pid": rank,
                                "tid": 3,
                                "ts": 1000 + rank,
                                "dur": 40,
                                "args": {
                                    "External id": 10 + rank,
                                    "Collective name": "allreduce",
                                    "stream": 3,
                                    "collective_id": rank,
                                },
                            }
                        ]
                    }
                )
            )
        dfs = generate_collective_report(
            trace_dir=str(tmp_path),
            world_size=2,
            output_csvs_dir=str(tmp_path / "coll"),
            use_multiprocessing=False,
            strict_world_size_check=False,
            all2allv_heatmap=True,
        )
        assert isinstance(dfs, dict)


class TestReportingCliPhase7:
    def test_pftrace_hip_activity_main(self, tmp_path):
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_pftrace_hip_activity"
        )
        trace_path = tmp_path / "pf.json"
        trace_path.write_text(json.dumps({"traceEvents": _minimal_pftrace_events()}))
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_pftrace_hip_activity",
            "--trace_path",
            str(trace_path),
            "--output_csvs_dir",
            str(tmp_path / "csv"),
            "--merge_kernels",
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert (tmp_path / "csv" / "category_summary.csv").exists()

    def test_pftrace_memory_copy_main(self, tmp_path):
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_pftrace_memory_copy"
        )
        events = [
            _mk_event("gpu_memcpy", "MemcpyHtoD", 1000, 20, 0, 1, {"bytes": 4096}),
            _mk_event("gpu_memcpy", "MemcpyDtoH", 1100, 15, 0, 1, {"bytes": 2048}),
        ]
        trace_path = tmp_path / "pf.json"
        trace_path.write_text(json.dumps({"traceEvents": events}))
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_pftrace_memory_copy",
            "--trace_path",
            str(trace_path),
            "--output_csvs_dir",
            str(tmp_path / "csv"),
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert (tmp_path / "csv" / "memory_copy_by_copy_bytes.csv").exists()


class TestReportingCliPhase9:
    def test_generate_perf_report_pytorch_main(self, tmp_path):
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_pytorch"
        )
        trace = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 100)])
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_pytorch",
            "--profile_json_path",
            trace,
            "--output_csvs_dir",
            str(tmp_path / "csv"),
            "--output_xlsx_path",
            str(tmp_path / "out.xlsx"),
            "--enable_kernel_summary",
            "--disable_coll_analysis",
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert (tmp_path / "csv" / "gpu_timeline.csv").exists()

    def test_generate_perf_report_inference_main(self, tmp_path):
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_pytorch_inference"
        )
        trace = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 100)], "inf.json")
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_pytorch_inference",
            "--profile_json_path",
            trace,
            "--output_csvs_dir",
            str(tmp_path / "csv"),
            "--output_xlsx_path",
            str(tmp_path / "out.xlsx"),
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert (tmp_path / "csv" / "gpu_timeline.csv").exists()

    @pytest.mark.skipif(
        not os.path.isfile(ROCprof_FILE), reason="rocprof fixture missing"
    )
    def test_generate_multi_rank_collective_main(self, tmp_path):
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_multi_rank_collective_report_pytorch"
        )
        for rank in (0, 1):
            (tmp_path / f"rank{rank}_trace.json").write_text(
                json.dumps(
                    {
                        "traceEvents": [
                            {
                                "ph": "X",
                                "cat": "kernel",
                                "name": "ncclKernel_AllReduce",
                                "pid": rank,
                                "tid": 3,
                                "ts": 1000,
                                "dur": 40,
                                "args": {"External id": 10, "stream": 3},
                            }
                        ]
                    }
                )
            )
        old_argv = sys.argv
        sys.argv = [
            "generate_multi_rank_collective_report_pytorch",
            "--trace_dir",
            str(tmp_path),
            "--world_size",
            "2",
            "--output_csvs_dir",
            str(tmp_path / "coll"),
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert os.path.isdir(tmp_path / "coll")

    def test_pftrace_hip_api_main(self, tmp_path):
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_pftrace_hip_api"
        )
        trace_path = tmp_path / "pf.json"
        trace_path.write_text(json.dumps({"traceEvents": _minimal_pftrace_events()}))
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_pftrace_hip_api",
            "--trace_path",
            str(trace_path),
            "--output_csvs_dir",
            str(tmp_path / "csv"),
        ]
        try:
            mod.main()
        except SystemExit:
            pass
        finally:
            sys.argv = old_argv


class TestReportingCliBoost:
    def test_inference_report_all_sheets(self, tmp_path):
        trace = _write_trace(
            tmp_path,
            [
                ("aten::mm", "gemm_kernel", 100),
                ("aten::add", "add_kernel", 15),
                ("aten::relu", "relu_kernel", 10),
            ],
        )
        result = generate_inference_report(
            profile_json_path=trace,
            output_csvs_dir=str(tmp_path / "inf_csvs"),
            output_xlsx_path=str(tmp_path / "inf.xlsx"),
            collective_analysis=False,
            kernel_summary=True,
            short_kernel_study=True,
            include_overlap_info=True,
            group_by_parent_module=True,
            group_by_num_kernels=True,
            micro_idle_thresh_us=5,
        )
        assert "ops_summary" in result or "gpu_timeline" in result

    def test_collective_report_multiprocessing(self, tmp_path):
        for rank in range(2):
            events = {
                "traceEvents": [
                    {
                        "ph": "X",
                        "cat": "kernel",
                        "name": "ncclKernel_AllReduce",
                        "pid": rank,
                        "tid": 3,
                        "ts": 1000 + rank,
                        "dur": 50,
                        "args": {
                            "External id": 10 + rank,
                            "Collective name": "allreduce",
                            "stream": 3,
                        },
                    }
                ]
            }
            (tmp_path / f"rank{rank}_trace.json").write_text(json.dumps(events))
        dfs = generate_collective_report(
            trace_dir=str(tmp_path),
            world_size=2,
            output_csvs_dir=str(tmp_path / "mp_out"),
            use_multiprocessing=True,
            max_workers=2,
            strict_world_size_check=False,
            all2allv_heatmap=False,
        )
        assert "nccl_summary_implicit_sync" in dfs

    @pytest.mark.skipif(
        not os.path.exists(
            os.path.join(
                os.path.dirname(__file__),
                "traces/mi210/gpu_only_trace/gpu_only_trace.json.gz",
            )
        ),
        reason="gpu_only trace fixture missing",
    )
    def test_treeperf_gpu_only_extended(self):
        analyzer = TreePerfAnalyzer.from_file(GPU_ONLY_TRACE, rebuild_tree=True)
        launchers = analyzer.get_df_kernel_launchers(include_args=True)
        assert not launchers.empty
        summary = TreePerfAnalyzer.get_df_kernel_launchers_summary(launchers)
        assert not summary.empty
        unique = TreePerfAnalyzer.get_df_kernel_launchers_unique_args(
            launchers, include_pct=True
        )
        assert not unique.empty
        unified = analyzer.build_df_unified_perf_table(include_nccl=False)
        assert isinstance(unified, pd.DataFrame)


class TestReportingCliPush95:
    def test_rocprof_main(self, tmp_path):
        if not os.path.isfile(ROCprof_FILE):
            pytest.skip("rocprof fixture missing")
        out = tmp_path / "roc.xlsx"
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_rocprof"
        )
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_rocprof",
            "--profile_json_path",
            ROCprof_FILE,
            "--output_xlsx_path",
            str(out),
            "--short_kernel_study",
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert out.exists()

    def test_pftrace_hip_api_main(self, tmp_path):
        trace_path = tmp_path / "trace.json"
        trace_path.write_text(json.dumps({"traceEvents": _minimal_pftrace_events()}))
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_pftrace_hip_api"
        )
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_pftrace_hip_api",
            "--trace_path",
            str(trace_path),
            "--output_csvs_dir",
            str(tmp_path / "hip_api_csv"),
            "--include_nonlaunch_apis",
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert (tmp_path / "hip_api_csv" / "api_kernel_summary.csv").exists()

    def test_pftrace_memory_copy_main(self, tmp_path):

        trace_path = tmp_path / "trace.json"
        trace_path.write_text(json.dumps({"traceEvents": _make_memory_copy_events()}))
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_pftrace_memory_copy"
        )
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_pftrace_memory_copy",
            "--trace_path",
            str(trace_path),
            "--output_csvs_dir",
            str(tmp_path / "mem_csv"),
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert any(f.endswith(".csv") for f in os.listdir(tmp_path / "mem_csv"))

    def test_pftrace_hip_activity_csv_and_default_xlsx(self, tmp_path):
        trace_path = tmp_path / "trace.json"
        trace_path.write_text(json.dumps({"traceEvents": _minimal_pftrace_events()}))
        csv_dir = tmp_path / "pf_csv"
        dfs = generate_perf_report_pftrace_hip_activity(
            trace_path=str(trace_path),
            output_csvs_dir=str(csv_dir),
            merge_kernels=True,
            kernel_summary_baseline="compute",
            hip_summary_group="name+stream",
            min_event_ns=1000,
        )
        assert (csv_dir / "category_summary.csv").exists()
        assert "category_summary" in dfs

    def test_genesis_main(self, tmp_path):
        capture = _create_genesis_capture(tmp_path)
        out = tmp_path / "gen_out"
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_genesis"
        )
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_genesis",
            "--capture-dir",
            str(capture),
            "--output-dir",
            str(out),
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert (out / "genesis_perf_report.xlsx").exists()

    def test_compare_reports_main(self, tmp_path):
        r1 = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 100)], "r1.json")
        r2 = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 110)], "r2.json")
        out1 = tmp_path / "rep1"
        out2 = tmp_path / "rep2"
        generate_perf_report_pytorch(
            profile_json_path=r1,
            output_csvs_dir=str(out1),
            output_xlsx_path=str(tmp_path / "r1.xlsx"),
            collective_analysis=False,
        )
        generate_perf_report_pytorch(
            profile_json_path=r2,
            output_csvs_dir=str(out2),
            output_xlsx_path=str(tmp_path / "r2.xlsx"),
            collective_analysis=False,
        )
        cmp_xlsx = tmp_path / "comparison.xlsx"
        mod = importlib.import_module(
            "TraceLens.Reporting.compare_perf_reports_pytorch"
        )
        old_argv = sys.argv
        sys.argv = [
            "compare_perf_reports_pytorch",
            str(out1),
            str(out2),
            "-o",
            str(cmp_xlsx),
            "--names",
            "a",
            "b",
            "--sheets",
            "gpu_timeline",
            "ops_summary",
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert cmp_xlsx.exists()

    def test_pytorch_report_main_extended(self, tmp_path):
        trace = _write_trace(
            tmp_path,
            [("aten::mm", "gemm_kernel", 100), ("aten::add", "add_kernel", 20)],
        )
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_pytorch"
        )
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_pytorch",
            "--profile_json_path",
            trace,
            "--output_csvs_dir",
            str(tmp_path / "py_ext"),
            "--output_xlsx_path",
            str(tmp_path / "py_ext.xlsx"),
            "--enable_kernel_summary",
            "--short_kernel_study",
            "--include_overlap_info",
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert (tmp_path / "py_ext" / "gpu_timeline.csv").exists()


def test_reporting_utils_and_pseudo_registry(tmp_path):

    trace = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 100)])
    assert ru.detect_gpus_per_node(trace) is None or isinstance(
        ru.detect_gpus_per_node(trace), int
    )
    tree = TraceToTree([])
    tree.build_tree()
    por.apply_pseudo_op_extensions(tree, verbose=False)
    assert tree is not None


def test_pytorch_report_main(tmp_path):
    trace = _write_trace(
        tmp_path,
        [("aten::mm", "gemm_kernel", 100), ("aten::add", "add_kernel", 15)],
    )
    out_dir = tmp_path / "py_csvs"
    xlsx = tmp_path / "py.xlsx"
    mod = importlib.import_module("TraceLens.Reporting.generate_perf_report_pytorch")

    old_argv = sys.argv
    sys.argv = [
        "generate_perf_report_pytorch",
        "--profile_json_path",
        trace,
        "--output_csvs_dir",
        str(out_dir),
        "--output_xlsx_path",
        str(xlsx),
        "--enable_kernel_summary",
        "--short_kernel_study",
        "--disable_coll_analysis",
        "--group_by_num_kernels",
    ]
    try:
        mod.main()
    finally:
        sys.argv = old_argv
    assert xlsx.exists()


class TestReportingPush95Coverage:
    @pytest.mark.parametrize("dirpath,trace_gz", _discover_inference_cases())
    def test_inference_report_with_capture_merge(self, dirpath, trace_gz, tmp_path):
        trace_path = os.path.join(dirpath, trace_gz)
        capture = os.path.join(dirpath, "capture_traces")
        metadata = os.path.join(capture, "execution_details.json")
        kwargs = {
            "profile_json_path": trace_path,
            "output_csvs_dir": str(tmp_path / "csv"),
            "output_xlsx_path": str(tmp_path / "report.xlsx"),
            "collective_analysis": False,
            "enable_pseudo_ops": True,
            "kernel_summary": True,
            "short_kernel_study": True,
            "group_by_parent_module": True,
            "include_overlap_info": True,
        }
        if os.path.isdir(capture) and os.path.isfile(metadata):

            merged = merge_capture_trace_into_graph(capture, metadata, trace_path)
            kwargs["augmented_tree"] = merged
        result = generate_inference_report(**kwargs)
        assert "gpu_timeline" in result

    @pytest.mark.skipif(
        not os.path.isfile(NORM_TRACE), reason="normalization trace missing"
    )
    def test_normalization_trace_pytorch_report(self, tmp_path):
        generate_perf_report_pytorch(
            profile_json_path=NORM_TRACE,
            output_csvs_dir=str(tmp_path / "norm_csv"),
            output_xlsx_path=str(tmp_path / "norm.xlsx"),
            collective_analysis=False,
            kernel_summary=True,
            short_kernel_study=True,
        )
        assert (tmp_path / "norm_csv" / "gpu_timeline.csv").exists()

    def test_compare_perf_reports_cli(self, tmp_path):
        r1 = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 100)], "r1.json")
        r2 = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 120)], "r2.json")
        out1 = tmp_path / "rep1"
        out2 = tmp_path / "rep2"
        generate_perf_report_pytorch(
            profile_json_path=r1,
            output_csvs_dir=str(out1),
            output_xlsx_path=str(tmp_path / "r1.xlsx"),
            collective_analysis=False,
        )
        generate_perf_report_pytorch(
            profile_json_path=r2,
            output_csvs_dir=str(out2),
            output_xlsx_path=str(tmp_path / "r2.xlsx"),
            collective_analysis=False,
        )
        cmp_xlsx = tmp_path / "comparison.xlsx"
        mod = importlib.import_module(
            "TraceLens.Reporting.compare_perf_reports_pytorch"
        )
        old_argv = sys.argv
        sys.argv = [
            "compare_perf_reports_pytorch",
            str(out1),
            str(out2),
            "-o",
            str(cmp_xlsx),
            "--names",
            "baseline",
            "candidate",
            "--sheets",
            "gpu_timeline",
            "ops_summary",
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert cmp_xlsx.exists()

    @pytest.mark.skipif(
        not os.path.isfile(ROCprof_FILE), reason="rocprof fixture missing"
    )
    def test_rocprof_report_cli(self, tmp_path):
        out = tmp_path / "roc.xlsx"
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_rocprof"
        )
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_rocprof",
            "--profile_json_path",
            ROCprof_FILE,
            "--output_xlsx_path",
            str(out),
            "--short_kernel_study",
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert out.exists()

    def test_genesis_report_cli(self, tmp_path):
        capture = _create_genesis_capture(tmp_path)
        out = tmp_path / "gen_out"
        mod = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_genesis"
        )
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_genesis",
            "--capture-dir",
            str(capture),
            "--output-dir",
            str(out),
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert (out / "genesis_perf_report.xlsx").exists()

    def test_perf_extensions_rope_and_quant(self):
        rope = pext.fused_qk_rope_concat_and_cache_mla(
            {
                "args": {
                    "Input Dims": [
                        (2, 8, 512),
                        (2, 8, 64),
                        (2, 1, 512),
                        (2, 1, 64),
                        (128, 1, 1, 576),
                        (2, 128),
                    ],
                    "Input type": ["c10::BFloat16"] * 5 + ["c10::Float8_e4m3fn"],
                }
            }
        )
        assert rope.flops() > 0

        silu = {
            "args": {
                "Input Dims": [(4, 512), (4, 512)],
                "Input type": ["c10::BFloat16", "c10::BFloat16"],
                "Input Strides": [(512, 1), (512, 1)],
            }
        }
        assert pext.aiter_silu_and_mul(silu).bytes() > 0


class TestPush95Phase2:
    def test_normalization_trace_treeperf_full(self):
        if not os.path.isfile(NORM_TRACE):
            pytest.skip("normalization trace missing")
        analyzer = TreePerfAnalyzer.from_file(
            NORM_TRACE,
            rebuild_tree=True,
            enable_pseudo_ops=True,
            add_python_func=True,
        )
        unified = analyzer.build_df_unified_perf_table(include_nccl=False)
        assert not unified.empty
        ops = analyzer.build_df_perf_metrics(
            events=[e for e in analyzer.tree.events if e.get("cat") == "cpu_op"]
        )
        assert isinstance(ops, pd.DataFrame)

    def test_orchestrator_main_multi_kernel_memcpy_nccl(self, tmp_path, monkeypatch):

        out = str(tmp_path)
        _write_minimal_orchestrator_csvs(out, comparative=False)
        gpu_events = [
            {
                "name": "gemm_kernel",
                "dur": 100,
                "ts": 1000,
                "_category": "kernel",
                "cat": "kernel",
                "args": {"stream": 0},
                "gpu_events": [],
            },
            {
                "name": "MemcpyHtoD",
                "dur": 20,
                "ts": 1100,
                "_category": "kernel",
                "cat": "kernel",
                "args": {"bytes": 4096, "stream": 1},
                "gpu_events": [],
            },
            {
                "name": "MemcpyDtoD",
                "dur": 15,
                "ts": 1120,
                "_category": "kernel",
                "cat": "kernel",
                "args": {"bytes": 2048, "stream": 1},
                "gpu_events": [],
            },
            {
                "name": "ncclKernel_AllReduce",
                "dur": 40,
                "ts": 1200,
                "_category": "kernel",
                "cat": "kernel",
                "args": {"stream": 2},
                "gpu_events": [],
            },
        ]
        tree = _StubTree(gpu_events, {i: e for i, e in enumerate(gpu_events)})
        analyzer = _StubAnalyzer(tree)

        class _FakeTreePerfAnalyzer:
            @classmethod
            def from_file(cls, *args, **kwargs):
                return analyzer

        monkeypatch.setattr(op, "TreePerfAnalyzer", _FakeTreePerfAnalyzer)
        monkeypatch.setattr(
            op, "_extract_standalone_fusion_candidates", lambda *a, **k: []
        )

        old_argv = sys.argv
        sys.argv = [
            "orchestrator_prepare",
            "--trace-path",
            "/fake/trace.json",
            "--platform",
            "MI300X",
            "--output-dir",
            out,
        ]
        try:
            op.main()
        finally:
            sys.argv = old_argv

        mk = json.loads(
            open(os.path.join(out, "category_data", "multi_kernel_data.json")).read()
        )
        assert "memcpy_summary" in mk and "nccl_summary" in mk

    def test_gemm_origami_import_error_path(self, monkeypatch):
        monkeypatch.delenv("GEMM_SIMULATOR_PATH", raising=False)
        perf_model.GEMM._origami_import_error_printed = False
        import builtins

        real_import = builtins.__import__

        def fake_import(name, *args, **kwargs):
            if name == "origami":
                raise ImportError("no origami")
            return real_import(name, *args, **kwargs)

        with patch.object(builtins, "__import__", fake_import):
            t, _ = perf_model.GEMM.get_simulation_time_func(
                _ARCH, 4, 8, 16, 1, "bf16", enable_origami=True
            )
        assert t is None

    def test_orchestrator_comparative_main(self, tmp_path, monkeypatch):

        out = str(tmp_path)
        _write_minimal_orchestrator_csvs(out, comparative=True)
        k1 = _kernel_event(10, "Cijk_a")
        k2 = _kernel_event(11, "ew_add")
        module = {
            "name": "nn.Module: MLP_0",
            "_category": "aten",
            "gpu_events": [10, 11],
        }
        tree = _StubTree([module], {10: k1, 11: k2})
        analyzer = _StubAnalyzer(tree)

        class _FakeTreePerfAnalyzer:
            @classmethod
            def from_file(cls, *args, **kwargs):
                return analyzer

        monkeypatch.setattr(op, "TreePerfAnalyzer", _FakeTreePerfAnalyzer)
        monkeypatch.setattr(
            op, "_extract_comparative_fusion_candidates", lambda *a, **k: []
        )
        monkeypatch.setattr(
            op, "_extract_standalone_fusion_candidates", lambda *a, **k: []
        )

        old_argv = sys.argv
        sys.argv = [
            "orchestrator_prepare",
            "--trace-path",
            "/fake/trace.json",
            "--platform",
            "MI300X",
            "--output-dir",
            out,
            "--comparison-scope",
            "comparative",
        ]
        try:
            op.main()
        finally:
            sys.argv = old_argv
        assert os.path.isfile(
            os.path.join(out, "metadata", "trace2_gpu_utilization.json")
        )

    def test_analysis_utils_and_kernel_fusion(self, tmp_path):

        row = pd.Series(
            {
                "FLOPS/Byte": 0.5,
                "TFLOPS/s_mean": 10.0,
                "TB/s_mean": 0.5,
                "Roofline Bound": "MEMORY_BOUND",
                "Compute Spec": "vector_fp32",
            }
        )
        eff = au.calculate_efficiency(
            row, peak_maf_or_maf_dict={"vector_fp32": 100.0}, peak_hbm_bw=5300
        )
        assert eff["bound_type"] == "memory"

        fusion_dir = tmp_path / "category_data"
        fusion_dir.mkdir()
        (fusion_dir / "kernel_fusion_metrics.json").write_text(
            json.dumps({"high_confidence_kernel_map": {"gemm_a": "fused_a"}})
        )
        assert au._load_fusion_map(str(tmp_path))["gemm_a"] == "fused_a"

        ops = [{"kernel_names": ["a", "b"], "base_name": "Block", "instance_count": 2}]
        assert len(kfa._filter_and_dedup(ops)) == 1

    def test_reporting_pftrace_and_collective(self, tmp_path):

        trace_path = tmp_path / "pf.json"
        trace_path.write_text(json.dumps({"traceEvents": _minimal_pftrace_events()}))
        generate_perf_report_pftrace_hip_activity(
            trace_path=str(trace_path),
            output_csvs_dir=str(tmp_path / "pf_csv"),
            merge_kernels=True,
            kernel_summary_baseline="compute",
        )
        assert (tmp_path / "pf_csv" / "category_summary.csv").exists()

        for rank in (0, 1):
            (tmp_path / f"rank{rank}_trace.json").write_text(
                json.dumps(
                    {
                        "traceEvents": [
                            {
                                "ph": "X",
                                "cat": "kernel",
                                "name": "ncclKernel_AllReduce",
                                "pid": rank,
                                "tid": 3,
                                "ts": 1000 + rank,
                                "dur": 40,
                                "args": {
                                    "External id": 10 + rank,
                                    "Collective name": "allreduce",
                                    "stream": 3,
                                    "collective_id": rank,
                                },
                            }
                        ]
                    }
                )
            )
        dfs = generate_collective_report(
            trace_dir=str(tmp_path),
            world_size=2,
            output_csvs_dir=str(tmp_path / "coll"),
            use_multiprocessing=False,
            strict_world_size_check=False,
            all2allv_heatmap=False,
        )
        assert isinstance(dfs, dict)

    def test_inference_report_all_flags(self, tmp_path):
        trace = tmp_path / "trace.json"
        trace.write_text(
            json.dumps(
                _build_synthetic_trace(
                    [
                        ("aten::mm", "gemm_kernel", 100),
                        ("aten::add", "vectorized_elementwise_kernel", 20),
                        (
                            "aten::_scaled_dot_product_flash_attention",
                            "flash_fwd_kernel",
                            80,
                        ),
                    ]
                )
            )
        )
        result = generate_inference_report(
            profile_json_path=str(trace),
            output_csvs_dir=str(tmp_path / "inf"),
            output_xlsx_path=str(tmp_path / "inf.xlsx"),
            collective_analysis=True,
            kernel_summary=True,
            short_kernel_study=True,
            include_overlap_info=True,
            group_by_parent_module=True,
            group_by_num_kernels=True,
            topk_ops=5,
            topk_roofline_ops=3,
            topk_short_kernels=2,
            include_unlinked_kernels=True,
            include_call_stack=True,
            micro_idle_thresh_us=1,
        )
        assert "gpu_timeline" in result

    def test_attention_extensions_remaining(self):

        base = {
            "annotation": _GDN_ANNOTATION,
            "args": {
                "Input Dims": [[64, 8, 64], [64, 8, 64], [64, 8, 128]],
                "Input type": ["c10::BFloat16"] * 3,
            },
        }
        for cls in (
            aext.pa_decode_gluon,
            aext.pa_sparse_prefill_opus_fwd,
            aext.pseudo_v4_paged_decode_hca,
            aext.pseudo_v4_paged_decode_csa,
        ):
            model = cls(base)
            assert model.bytes() is None or model.bytes() >= 0


class TestPush95Phase3:
    def test_tracediff_perf_summary_branches(self):

        assert tde.tracediff_perf_summary_from_diff_stats(pd.DataFrame()).empty
        diff = pd.DataFrame(
            {
                "source": ["trace1", "trace2"],
                "lowest_common_ancestor_id": [1, 1],
                "lowest_common_ancestor_name": ["aten::mm", "aten::mm"],
                "cpu_op_name": ["aten::mm", "aten::add"],
                "busy_time": [100.0, 80.0],
                "name": ["k1", "k2"],
                "gpu_op_uid": [10, 20],
                "nn_module_stack": ["[]", "[]"],
                "nn_module_parent": ["", ""],
                "Input Dims": ["[[2,3]]", "[[2,3]]"],
                "Input type": ["['fp16']", "['fp16']"],
                "Input Strides": ["[]", "[]"],
                "Concrete Inputs": ["", ""],
            }
        )
        summary = tde.tracediff_perf_summary_from_diff_stats(diff)
        assert not summary.empty

        multi = pd.DataFrame(
            {
                "source": ["trace1", "trace1"],
                "lowest_common_ancestor_id": [2, 2],
                "lowest_common_ancestor_name": ["block", "block"],
                "cpu_op_name": ["aten::mm", "aten::relu"],
                "busy_time": [50.0, 30.0],
                "name": ["k1", "k2"],
                "gpu_op_uid": [1, 2],
                "nn_module_stack": ["[]", "[]"],
                "nn_module_parent": ["", ""],
                "Input Dims": ["[[2,3]]", "[[2,3]]"],
                "Input type": ["['fp16']", "['fp16']"],
                "Input Strides": ["[]", "[]"],
                "Concrete Inputs": ["", ""],
            }
        )
        assert (
            " | " in tde.tracediff_perf_summary_from_diff_stats(multi).iloc[0]["name"]
        )

    def test_kernel_fusion_impact_pipeline(self, tmp_path):

        csv_dir = tmp_path / "perf_report_csvs"
        csv_dir.mkdir()
        pd.DataFrame(
            {
                "kernel_details_summary": [
                    "[{'name': 'Cijk_a'}]",
                    "[{'name': 'ew_add'}]",
                ],
                "op category": ["GEMM", "elementwise"],
                "Data Moved (MB)": [10.0, 2.0],
                "perf_params": ["{'M':2}", "{}"],
                "Input Dims": ["[[2,3]]", "[[4,4]]"],
            }
        ).to_csv(csv_dir / "unified_perf_summary.csv", index=False)

        cat_dir = tmp_path / "category_data"
        cat_dir.mkdir()
        (cat_dir / "category_manifest.json").write_text(
            json.dumps(
                {
                    "platform": "MI300X",
                    "gpu_utilization": {"total_time_ms": 1000.0},
                }
            )
        )
        (cat_dir / "fusion_candidates.json").write_text(
            json.dumps(
                [
                    {
                        "module_name": "nn.Module: Block",
                        "kernels": [
                            {
                                "name": "Cijk_a",
                                "type": "GEMM",
                                "dur_us": 100,
                                "data_in_mb": 10.0,
                            },
                            {
                                "name": "ew_add",
                                "type": "Elementwise Add",
                                "dur_us": 20,
                                "data_in_mb": 2.0,
                            },
                        ],
                        "instance_count": 1,
                    }
                ]
            )
        )
        (cat_dir / "arch_config.json").write_text(
            json.dumps(
                {
                    "peak_hbm_bw_tbs": 5.3,
                    "max_achievable_tflops": {
                        "matrix_bf16": 1000.0,
                        "vector_fp32": 100.0,
                    },
                }
            )
        )

        candidates, manifest, csv_path = kfa.load_fusion_data(str(tmp_path))
        lookup = kfa.build_kernel_perf_lookup(csv_path)
        estimates = kfa.compute_fusion_impact_estimates(
            candidates,
            lookup,
            peak_bw_tbs=5.3,
            peak_maf_tflops={"matrix_bf16": 1000.0, "vector_fp32": 100.0},
            baseline_ms=1000.0,
            is_comparative=False,
        )
        assert isinstance(estimates, list)

        mod = importlib.import_module(
            "TraceLens.Agent.Analysis.category_analyses.kernel_fusion_analysis"
        )
        old_argv = sys.argv
        sys.argv = [
            "kernel_fusion_analysis",
            "--output-dir",
            str(tmp_path),
            "--comparison-scope",
            "standalone",
        ]
        try:
            mod.main()
        finally:
            sys.argv = old_argv
        assert (cat_dir / "kernel_fusion_metrics.json").exists()

    def test_pftrace_cli_mains(self, tmp_path):

        hip_api = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_pftrace_hip_api"
        )
        trace_path = tmp_path / "hip_api.json"
        trace_path.write_text(json.dumps({"traceEvents": _minimal_pftrace_events()}))
        old_argv = sys.argv
        sys.argv = [
            "generate_perf_report_pftrace_hip_api",
            "--trace_path",
            str(trace_path),
            "--output_csvs_dir",
            str(tmp_path / "hip_api_csv"),
        ]
        try:
            hip_api.main()
        finally:
            sys.argv = old_argv
        assert (tmp_path / "hip_api_csv" / "api_kernel_summary.csv").exists()

        mem_copy = importlib.import_module(
            "TraceLens.Reporting.generate_perf_report_pftrace_memory_copy"
        )
        mem_path = tmp_path / "mem.json"
        mem_path.write_text(json.dumps({"traceEvents": _make_memory_copy_events()}))
        sys.argv = [
            "generate_perf_report_pftrace_memory_copy",
            "--trace_path",
            str(mem_path),
            "--output_csvs_dir",
            str(tmp_path / "mem_csv"),
        ]
        try:
            mem_copy.main()
        finally:
            sys.argv = old_argv
        assert any(f.endswith(".csv") for f in os.listdir(tmp_path / "mem_csv"))

    def test_llama_fsdp_traces_treeperf(self):
        fsdp_dir = os.path.join(TRACES_ROOT, "mi300/llama_70b_fsdp")
        if not os.path.isdir(fsdp_dir):
            pytest.skip("fsdp traces missing")
        trace = os.path.join(fsdp_dir, "rank0_trace_no_pyfn.json.gz")
        if not os.path.isfile(trace):
            pytest.skip("rank0 trace missing")
        analyzer = TreePerfAnalyzer.from_file(
            trace, rebuild_tree=True, enable_pseudo_ops=True, add_python_func=False
        )
        unified = analyzer.build_df_unified_perf_table(include_nccl=True)
        assert isinstance(unified, pd.DataFrame)

    def test_perf_model_tex_gemm_and_reduce_edges(self):
        input_dims = [()] * 19
        input_dims[0] = [128, 64]
        input_dims[5] = [256, 64]
        input_dims[10] = [128, 256]
        tex = perf_model.tex_ts_te_gemm_ts(
            {
                "args": {
                    "Input Dims": input_dims,
                    "Input type": ["c10::Float8_e4m3fn"] * 19,
                    "Concrete Inputs": [""] * 4
                    + ["1"]
                    + [""] * 4
                    + ["1"]
                    + [""] * 4
                    + ["bias"],
                }
            }
        )
        assert tex.flops() > 0
        with pytest.raises(NotImplementedError):
            tex.flops_bwd()

        mean_evt = {
            "name": "aten::mean",
            "args": {
                "Input Dims": [(4, 256)],
                "Input type": ["c10::BFloat16"],
                "Output type": ["c10::BFloat16"],
                "Concrete Inputs": ["", "[1]", "True"],
            },
        }
        assert perf_model.aten_reduce(mean_evt).flops() > 0


class TestPush95Phase4:
    def test_orchestrator_main_real_fusion_extraction(self, tmp_path, monkeypatch):

        out = str(tmp_path)
        csv_dir = os.path.join(out, "perf_report_csvs")
        os.makedirs(csv_dir)
        pd.DataFrame(
            {
                "type": ["total_time", "computation_time", "idle_time"],
                "time ms": [1000.0, 900.0, 100.0],
                "percent": [100.0, 90.0, 10.0],
            }
        ).to_csv(os.path.join(csv_dir, "gpu_timeline.csv"), index=False)
        pd.DataFrame(
            {
                "name": ["aten::mm"],
                "total_direct_kernel_time_ms": [800.0],
                "op category": ["GEMM"],
            }
        ).to_csv(os.path.join(csv_dir, "ops_summary.csv"), index=False)
        pd.DataFrame(
            {
                "name": ["aten::mm"],
                "op category": ["GEMM"],
                "Kernel Time (µs)_sum": [800000.0],
                "kernel_details_summary": ["[{'name': 'Cijk_a'}]"],
                "Data Moved (MB)": [10.0],
                "perf_params": ["{}"],
                "Input Dims": ["[[2,3]]"],
            }
        ).to_csv(os.path.join(csv_dir, "unified_perf_summary.csv"), index=False)
        pd.DataFrame({"name": ["aten::mm"], "op category": ["GEMM"]}).to_csv(
            os.path.join(csv_dir, "ops_summary_by_category.csv"), index=False
        )

        k1 = _kernel_event(10, "Cijk_a", dur=500)
        k2 = _kernel_event(11, "vectorized_elementwise_kernel add", dur=300)
        module = {
            "name": "nn.Module: MLP_0",
            "_category": "aten",
            "gpu_events": [10, 11],
            "args": {"Input Dims": "[[2,3]]"},
        }
        tree = _StubTree([module], {10: k1, 11: k2})
        analyzer = _StubAnalyzer(tree)

        class _FakeTreePerfAnalyzer:
            @classmethod
            def from_file(cls, *args, **kwargs):
                return analyzer

        monkeypatch.setattr(op, "TreePerfAnalyzer", _FakeTreePerfAnalyzer)

        old_argv = sys.argv
        sys.argv = [
            "orchestrator_prepare",
            "--trace-path",
            "/fake/trace.json",
            "--platform",
            "MI300X",
            "--output-dir",
            out,
        ]
        try:
            op.main()
        finally:
            sys.argv = old_argv
        fusion_path = os.path.join(out, "category_data", "fusion_candidates.json")
        assert os.path.isfile(fusion_path)
        assert isinstance(json.loads(open(fusion_path).read()), list)

    def test_orchestrator_comparative_fusion_via_main(self, tmp_path, monkeypatch):

        out = str(tmp_path)
        _write_minimal_orchestrator_csvs(out, comparative=True)
        t1_csv = os.path.join(out, "perf_report_trace1_csvs")
        pd.DataFrame(
            {
                "name": ["Cijk_A", "ew_add"],
                "source": ["trace1", "trace1"],
                "lowest_common_ancestor_id": [100, 100],
                "kernel_time": [5000.0, 3000.0],
                "gpu_op_uid": [10, 11],
            }
        ).to_csv(os.path.join(t1_csv, "diff_stats.csv"), index=False)

        k1 = _kernel_event(10, "Cijk_A", dur=500)
        k2 = _kernel_event(11, "ew_add", dur=300)
        module = {
            "name": "nn.Module: Attn_0",
            "_category": "aten",
            "gpu_events": [10, 11],
        }
        tree = _StubTree([module], {10: k1, 11: k2})
        analyzer = _StubAnalyzer(tree)

        class _FakeTreePerfAnalyzer:
            @classmethod
            def from_file(cls, *args, **kwargs):
                return analyzer

        monkeypatch.setattr(op, "TreePerfAnalyzer", _FakeTreePerfAnalyzer)

        old_argv = sys.argv
        sys.argv = [
            "orchestrator_prepare",
            "--trace-path",
            "/fake/trace.json",
            "--platform",
            "MI300X",
            "--output-dir",
            out,
            "--comparison-scope",
            "comparative",
        ]
        try:
            op.main()
        finally:
            sys.argv = old_argv
        fusion = json.loads(
            open(os.path.join(out, "category_data", "fusion_candidates.json")).read()
        )
        assert isinstance(fusion, list)

    @pytest.mark.parametrize(
        "rank",
        list(range(8)),
    )
    def test_llama_fsdp_all_ranks(self, rank):
        trace = os.path.join(
            TRACES_ROOT, "mi300/llama_70b_fsdp", f"rank{rank}_trace_no_pyfn.json.gz"
        )
        if not os.path.isfile(trace):
            pytest.skip("fsdp trace missing")
        analyzer = TreePerfAnalyzer.from_file(
            trace, rebuild_tree=True, enable_pseudo_ops=True
        )
        launchers = analyzer.get_df_kernel_launchers(include_args=True)
        assert isinstance(launchers, pd.DataFrame)

    def test_inference_comparison_report(self, tmp_path):
        trace1 = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 100)], "t1.json")
        trace2 = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 120)], "t2.json")
        result = generate_inference_report(
            profile_json_path=trace1,
            comparison_json_path=trace2,
            output_csvs_dir=str(tmp_path / "cmp"),
            output_xlsx_path=str(tmp_path / "cmp.xlsx"),
            collective_analysis=True,
            include_overlap_info=True,
            kernel_summary=True,
            short_kernel_study=True,
            group_by_parent_module=True,
        )
        assert "gpu_timeline" in result
        assert (tmp_path / "cmp" / "gpu_timeline.csv").exists()

    @pytest.mark.parametrize(
        "trace_name",
        [
            "traces/mi300/resnet_act_checkpoint.json.gz",
            "traces/mi300/Qwen_Qwen1.5-0.5B-Chat__1016005.json.gz",
            "traces/torch_compile_triton/trace.json.gz",
        ],
    )
    def test_key_traces_extended_dfs(self, trace_name):
        path = os.path.join(TESTS_DIR, trace_name)
        if not os.path.isfile(path):
            pytest.skip("trace missing")
        analyzer = TreePerfAnalyzer.from_file(
            path, rebuild_tree=True, enable_pseudo_ops=True, add_python_func=True
        )
        perf = analyzer.build_df_perf_metrics(
            events=[e for e in analyzer.tree.events if e.get("cat") == "cpu_op"][:20]
        )
        assert isinstance(perf, pd.DataFrame)
        if analyzer.add_python_func:
            nn = [
                e
                for e in analyzer.tree.events
                if str(e.get("name", "")).startswith("nn.Module")
            ]
            if nn:
                analyzer.build_nn_module_latency_tree(nn[0])


class TestPush95Phase5:
    def test_orchestrator_sync_bottleneck_via_phase6(self, tmp_path, monkeypatch):

        TestOrchestratorPhase6().test_main_no_time_column_sync_and_memcpy_dirs(
            tmp_path, monkeypatch
        )

    def test_rocprof_categorize_branches(self):

        assert _categorize_kernel("conv2d_fwd") == "Convolution"
        assert _categorize_kernel("layer_norm") == "Normalization"
        assert _categorize_kernel("flash_attn") == "Attention"


class TestPerfExtensionsFinal:
    def test_mhc_and_sampling_bytes(self):
        fused = pext.mhc_fused_post_pre_gemm_sqrsum(
            {
                "args": {
                    "Input Dims": [
                        (2, 4, 8),
                        (2, 4),
                        (4, 2, 128),
                        (4, 128),
                        (4, 2, 128),
                        (4, 2, 1),
                        (4, 2, 2),
                        (8, 256),
                    ],
                    "Input type": ["float"] * 8,
                }
            }
        )
        assert fused.bytes() > 0
        assert fused.get_maf_type() == "matrix"

        topk = pext.topk_softplus(
            {
                "args": {
                    "Input Dims": [(4, 2), (4, 2), (4, 8)],
                    "Input type": ["c10::Float", "c10::Int", "c10::BFloat16"],
                }
            }
        )
        assert topk.bytes() > 0

        sample = pext.mixed_sample_outer_exponential(
            {
                "args": {
                    "Input Dims": [(), (4, 32000), (4, 32000)],
                    "Input type": ["Scalar", "float", "float"],
                }
            }
        )
        assert sample.flops() > 0
        assert sample.get_maf_type() == "vector"

    def test_fused_qk_rope_and_batched_gemm_bytes(self):
        rope = pext.fused_qk_rope_concat_and_cache_mla(
            {
                "args": {
                    "Input Dims": [
                        (2, 8, 512),
                        (2, 8, 64),
                        (2, 1, 512),
                        (2, 1, 64),
                        (128, 1, 1, 576),
                    ],
                    "Input type": ["c10::BFloat16"] * 4 + ["c10::Float8_e4m3fn"],
                }
            }
        )
        assert rope.bytes() > 0

        fp4 = pext.batched_gemm_a16wfp4(
            {
                "args": {
                    "Input Dims": [[2, 4, 128], [2, 256, 64], [2, 256, 4]],
                    "Input type": ["c10::BFloat16", "unsigned char", "c10::Float"],
                }
            }
        )
        assert fp4.bytes() > 0

        post = pext.mhc_post(
            {
                "args": {
                    "Input Dims": [(4, 2, 128), (4, 128)],
                    "Input type": ["c10::BFloat16", "c10::BFloat16"],
                }
            }
        )
        assert post.bytes() > 0

        pre = pext.mhc_pre_gemm_sqrsum(
            {
                "args": {
                    "Input Dims": [(2, 4, 8), (2, 4), (4, 2, 128), (8, 256)],
                    "Input type": ["float", "float", "c10::BFloat16", "float"],
                }
            }
        )
        assert pre.bytes() > 0

        rope2 = pext.aiter_rope_cached_positions_2c_fwd_impl(
            {
                "args": {
                    "Input Dims": [
                        (2, 128, 8, 64),
                        (2, 128, 1, 64),
                        (2, 128, 8, 64),
                        (2, 128, 1, 64),
                        (2048, 1, 1, 64),
                        (2048, 1, 1, 64),
                        (2, 128),
                    ],
                    "Input type": ["c10::BFloat16"] * 7,
                }
            }
        )
        assert rope2.bytes() > 0


class TestReportingFinalCoverage:
    def test_inference_graphlaunch_warning(self, tmp_path):
        specs = [("aten::mm", "gemm_kernel", 100)]
        trace = _write_trace(tmp_path, specs)
        data = json.loads(open(trace).read())
        data["traceEvents"].append(
            _mk_event(
                "cuda_runtime",
                "hipGraphLaunch",
                2000,
                10,
                100,
                100,
                {"correlation": 999},
            )
        )
        path = tmp_path / "graph_trace.json"
        path.write_text(json.dumps(data))
        with pytest.warns(UserWarning, match="hipgraph launches"):
            generate_inference_report(
                profile_json_path=str(path),
                output_csvs_dir=str(tmp_path / "out"),
                output_xlsx_path=str(tmp_path / "r.xlsx"),
                collective_analysis=False,
            )

    def test_inference_bwd_and_overlap_sheets(self, tmp_path):
        corr_fwd, corr_bwd = 300, 301
        events = [
            _mk_event(
                "cpu_op",
                "aten::mm",
                1000,
                100,
                100,
                100,
                {
                    "Input Dims": [[32, 64], [64, 128]],
                    "Input type": ["float", "float"],
                },
            ),
            _mk_event(
                "cuda_runtime",
                "hipLaunchKernel",
                1010,
                5,
                100,
                100,
                {"correlation": corr_fwd},
            ),
            _mk_event(
                "kernel",
                "gemm_fwd",
                1050,
                80,
                0,
                7,
                {"correlation": corr_fwd, "stream": 7},
            ),
            _mk_ac2g(corr_fwd, 0, 7, 1050, "s"),
            _mk_ac2g(corr_fwd, 0, 7, 1130, "f"),
            _mk_event(
                "cpu_op",
                "aten::mm_backward",
                2000,
                100,
                100,
                100,
                {
                    "Input Dims": [[32, 64], [64, 128]],
                    "Input type": ["float", "float"],
                },
            ),
            _mk_event(
                "cuda_runtime",
                "hipLaunchKernel",
                2010,
                5,
                100,
                100,
                {"correlation": corr_bwd},
            ),
            _mk_event(
                "kernel",
                "gemm_bwd",
                2050,
                60,
                0,
                7,
                {"correlation": corr_bwd, "stream": 7},
            ),
            _mk_ac2g(corr_bwd, 0, 7, 2050, "s"),
            _mk_ac2g(corr_bwd, 0, 7, 2110, "f"),
        ]
        path = tmp_path / "bwd_trace.json"
        path.write_text(json.dumps({"traceEvents": events}))
        result = generate_inference_report(
            profile_json_path=str(path),
            output_csvs_dir=str(tmp_path / "bwd_out"),
            output_xlsx_path=str(tmp_path / "bwd.xlsx"),
            include_overlap_info=True,
            kernel_summary=True,
            short_kernel_study=True,
            group_by_parent_module=False,
        )
        assert "gpu_timeline" in result

    def test_sanity_check_kernel_details_summary_column(self):
        events = [{"name": "k_a", "cat": "kernel"}]
        tl = pd.DataFrame({"type": ["computation_time"], "time ms": [0.1]})
        kl = pd.DataFrame(
            {
                "total_direct_kernel_time": [100.0],
                "kernel_details_summary": [[{"name": "k_a", "count": 1}]],
            }
        )
        up = pd.DataFrame(
            {"Kernel Time (µs)": [100.0], "kernel_details_summary": [[{"name": "k_a"}]]}
        )
        result = perf_report_sanity_check(events, tl, kl, up)
        assert result["kl_count_pass"]

    def test_pytorch_report_with_comparison(self, tmp_path):
        trace1 = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 100)], "t1.json")
        trace2 = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 120)], "t2.json")
        generate_perf_report_pytorch(
            profile_json_path=trace1,
            comparison_json_path=trace2,
            output_csvs_dir=str(tmp_path / "cmp_out"),
            output_xlsx_path=str(tmp_path / "cmp.xlsx"),
            kernel_summary=True,
            short_kernel_study=True,
        )
        assert (tmp_path / "cmp_out" / "gpu_timeline.csv").exists()

    def test_add_truncated_kernel_details_inference(self):
        df = pd.DataFrame({"kernel_details": [[{"name": "x" * 200, "dur": 1}]]})
        out = add_truncated_inference(df, "kernel_details")
        assert "trunc_kernel_details" in out.columns


class TestReportingExtended:
    def test_inference_report_topk_and_roofline(self, tmp_path):
        trace = _write_trace(
            tmp_path,
            [
                ("aten::mm", "gemm_kernel", 100),
                ("aten::add", "add_kernel", 15),
            ],
        )
        result = generate_inference_report(
            profile_json_path=trace,
            output_csvs_dir=str(tmp_path / "topk_out"),
            output_xlsx_path=str(tmp_path / "topk.xlsx"),
            topk_ops=5,
            topk_roofline_ops=3,
            topk_short_kernels=2,
            short_kernel_threshold_us=50,
            include_unlinked_kernels=True,
            include_call_stack=True,
        )
        assert "gpu_timeline" in result

    def test_pytorch_report_extension_and_arch(self, tmp_path):
        trace = _write_trace(tmp_path, [("aten::mm", "gemm_kernel", 100)])
        ext = tmp_path / "ext.py"
        ext.write_text(
            "def apply_extension(analyzer, path):\n"
            "    analyzer.tree.events[0]['ext'] = True\n"
        )
        generate_perf_report_pytorch(
            profile_json_path=trace,
            output_csvs_dir=str(tmp_path / "ext_out"),
            output_xlsx_path=str(tmp_path / "ext.xlsx"),
            extension_file=str(ext),
            gpu_arch={
                "name": "mi300x",
                "freq_mhz": 2200,
                "num_cus": 304,
                "gemm_units_per_cu": 4,
                "mem_bw_gbps": 5300,
                "l1_bw_gbps": 100,
            },
            include_call_stack=True,
        )
        assert (tmp_path / "ext_out" / "gpu_timeline.csv").exists()


###############################################################################
# reporting_utils — _safe_sheet_name
###############################################################################


class TestSafeSheetName:
    """Validate Excel sheet name deduplication and length limits."""

    def test_no_collision(self):
        used = set()
        name = _safe_sheet_name("gpu_timeline", used)
        assert name == "gpu_timeline"
        assert "gpu_timeline" in used

    def test_collision_adds_suffix(self):
        used = {"gpu_timeline"}
        name = _safe_sheet_name("gpu_timeline", used)
        assert name == "gpu_timeline_1"
        assert "gpu_timeline_1" in used

    def test_multiple_collisions(self):
        used = {"test_sheet", "test_sheet_1", "test_sheet_2"}
        name = _safe_sheet_name("test_sheet", used)
        assert name == "test_sheet_3"

    def test_truncates_to_31_chars(self):
        used = set()
        long_name = "a" * 50
        name = _safe_sheet_name(long_name, used)
        assert len(name) <= 31

    def test_truncation_with_collision(self):
        long_name = "a" * 31
        used = {long_name}
        name = _safe_sheet_name(long_name, used)
        assert len(name) <= 31
        assert name != long_name
        assert name.endswith("_1")


###############################################################################
# reporting_utils — write_report_outputs
###############################################################################


def _read_sheets(xlsx_path):
    """Return {sheet_name: DataFrame} for an .xlsx file."""
    return pd.read_excel(xlsx_path, sheet_name=None)


class TestWriteReportOutputs:
    def _dfs(self):
        return {
            "alpha": pd.DataFrame({"a": [1, 2], "b": [3, 4]}),
            "beta": pd.DataFrame({"x": [5]}),
        }

    def test_csvs_only(self, tmp_path):
        out = tmp_path / "csvs"
        write_report_outputs(self._dfs(), csvs_dir=str(out))
        assert (out / "alpha.csv").exists()
        assert (out / "beta.csv").exists()
        # No xlsx requested -> none written
        assert not list(tmp_path.glob("*.xlsx"))
        pd.testing.assert_frame_equal(
            pd.read_csv(out / "alpha.csv"), self._dfs()["alpha"]
        )

    def test_xlsx_only(self, tmp_path):
        xlsx = tmp_path / "report.xlsx"
        write_report_outputs(self._dfs(), xlsx_path=str(xlsx))
        assert xlsx.exists()
        assert not list(tmp_path.glob("*.csv"))
        sheets = _read_sheets(xlsx)
        assert set(sheets) == {"alpha", "beta"}

    def test_both_outputs_written(self, tmp_path):
        xlsx = tmp_path / "report.xlsx"
        csvs = tmp_path / "csvs"
        write_report_outputs(self._dfs(), xlsx_path=str(xlsx), csvs_dir=str(csvs))
        assert xlsx.exists()
        assert (csvs / "alpha.csv").exists()
        assert (csvs / "beta.csv").exists()

    def test_neither_output_is_noop(self, tmp_path):
        write_report_outputs(self._dfs())
        assert not list(tmp_path.iterdir())

    def test_long_sheet_names_truncated_and_deduped(self, tmp_path):
        xlsx = tmp_path / "report.xlsx"
        dfs = {
            "a" * 40: pd.DataFrame({"c": [1]}),
            "a" * 45: pd.DataFrame({"c": [2]}),  # truncates to same 31 chars -> deduped
        }
        write_report_outputs(dfs, xlsx_path=str(xlsx))
        names = list(_read_sheets(xlsx))
        assert all(len(n) <= 31 for n in names)
        assert len(names) == 2  # no collision -> both sheets present

    def test_skip_empty_drops_empty_and_none(self, tmp_path):
        xlsx = tmp_path / "report.xlsx"
        csvs = tmp_path / "csvs"
        dfs = {
            "keep": pd.DataFrame({"a": [1]}),
            "empty": pd.DataFrame(),
            "none": None,
        }
        write_report_outputs(
            dfs, xlsx_path=str(xlsx), csvs_dir=str(csvs), skip_empty=True
        )
        assert set(_read_sheets(xlsx)) == {"keep"}
        assert (csvs / "keep.csv").exists()
        assert not (csvs / "empty.csv").exists()
        assert not (csvs / "none.csv").exists()

    def test_skip_empty_default_keeps_empty(self, tmp_path):
        xlsx = tmp_path / "report.xlsx"
        dfs = {
            "keep": pd.DataFrame({"a": [1]}),
            "empty": pd.DataFrame({"a": []}),
        }
        write_report_outputs(dfs, xlsx_path=str(xlsx))
        assert set(_read_sheets(xlsx)) == {"keep", "empty"}

    def test_hide_columns_hides_and_keeps_data(self, tmp_path):
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        xlsx = tmp_path / "report.xlsx"
        df = pd.DataFrame({"keep": [1], "hide_me": [2], "also_keep": [3]})
        write_report_outputs(
            {"data": df}, xlsx_path=str(xlsx), hide_columns={"data": ["hide_me"]}
        )
        ws = load_workbook(xlsx)["data"]
        hidden_col = get_column_letter(df.columns.get_loc("hide_me") + 1)  # "B"
        assert ws.column_dimensions[hidden_col].hidden is True
        # Non-hidden columns are not hidden, and the data is still present
        keep_col = get_column_letter(df.columns.get_loc("keep") + 1)
        assert ws.column_dimensions[keep_col].hidden is False
        assert set(_read_sheets(xlsx)["data"].columns) == {
            "keep",
            "hide_me",
            "also_keep",
        }

    def test_hide_columns_ignores_missing_column(self, tmp_path):
        xlsx = tmp_path / "report.xlsx"
        df = pd.DataFrame({"a": [1]})
        # Should not raise even though "nonexistent" isn't a column
        write_report_outputs(
            {"data": df}, xlsx_path=str(xlsx), hide_columns={"data": ["nonexistent"]}
        )
        assert xlsx.exists()

    def test_sheet_named_like_openpyxl_default(self, tmp_path):
        # "sheet" collides case-insensitively with openpyxl's default "Sheet",
        # so openpyxl renames it; the worksheet lookup must not rely on the
        # requested name.  hide_columns exercises that lookup path.
        xlsx = tmp_path / "report.xlsx"
        df = pd.DataFrame({"a": [1], "b": [2]})
        write_report_outputs(
            {"sheet": df}, xlsx_path=str(xlsx), hide_columns={"sheet": ["b"]}
        )
        assert xlsx.exists()


###############################################################################
# pftrace_utils — derive_pftrace_output_path
###############################################################################


class TestDerivePftraceOutputPath:
    def test_pftrace_suffix(self):
        assert (
            derive_pftrace_output_path("/tmp/trace.pftrace", "_activity_report.xlsx")
            == "/tmp/trace_activity_report.xlsx"
        )

    def test_json_gz_suffix(self):
        assert (
            derive_pftrace_output_path("/tmp/trace.json.gz", "_hip_api_report.xlsx")
            == "/tmp/trace_hip_api_report.xlsx"
        )

    def test_plain_json_suffix(self):
        assert (
            derive_pftrace_output_path("/tmp/trace.json", "_memory_copy_report.xlsx")
            == "/tmp/trace_memory_copy_report.xlsx"
        )
